#!/usr/bin/env python3
"""
Control Financiero 2026 - Generador de Excel
Versión 3.0 - Equivalente al Google Apps Script
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date
from decimal import Decimal

# ============================================================================
# CONFIGURACIÓN - DATOS MAESTROS
# ============================================================================

CONFIG = {
    'ENTIDADES': ['FAMILIA', 'NEUROTEA'],
    'CUENTAS_FAMILIA': ['BCP_FAM', 'YAPE_FAM', 'EFECTIVO_FAM', 'COLCHON_FAM'],
    'CUENTAS_NEUROTEA': ['BCP_NT', 'YAPE_NT', 'EFECTIVO_NT'],
    'TIPOS_MOVIMIENTO': ['INGRESO', 'EGRESO', 'TRANSFERENCIA'],
    'ESTADOS': ['PENDIENTE', 'PAGADO', 'ANULADO'],
    'FRECUENCIAS': ['MENSUAL', 'QUINCENAL', 'SEMANAL', 'EVENTUAL', 'ANUAL'],
    'MESES': ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC'],
}

# Colores
COLORES = {
    'VERDE_FAMILIA': '059669',
    'VERDE_CLARO': 'D1FAE5',
    'AZUL_NEUROTEA': '1D4ED8',
    'AZUL_CLARO': 'DBEAFE',
    'GRIS_HEADER': '374151',
    'GRIS_CLARO': 'F3F4F6',
    'BLANCO': 'FFFFFF',
    'ROJO': 'DC2626',
    'ROJO_CLARO': 'FEE2E2',
    'AMARILLO': 'F59E0B',
    'AMARILLO_CLARO': 'FEF3C7',
    'VERDE_OK': '10B981',
}

# ============================================================================
# DATOS DEL PRESUPUESTO
# ============================================================================

DATOS = {
    'FAMILIA': {
        'INGRESOS_FIJOS': [
            {'ITEM': 'Sueldo Marco (de NT)', 'MONTO_BASE': 4000, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1},
            {'ITEM': 'Sueldo Pierina (de NT)', 'MONTO_BASE': 2500, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1},
            {'ITEM': 'Ingreso Fijo Reserva 1', 'MONTO_BASE': 0, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1},
            {'ITEM': 'Ingreso Fijo Reserva 2', 'MONTO_BASE': 0, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1},
            {'ITEM': 'Ingreso Fijo Reserva 3', 'MONTO_BASE': 0, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1},
        ],
        'INGRESOS_VARIABLES': [
            {'ITEM': 'Devolución préstamo de NT', 'CATEGORIA': 'PRESTAMO', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Otros ingresos', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Ingreso Variable Reserva 1', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Ingreso Variable Reserva 2', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Ingreso Variable Reserva 3', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
        ],
        'GASTOS_FIJOS': [
            {'ITEM': 'Alquiler', 'MONTO_BASE': 1600, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'VIVIENDA'},
            {'ITEM': 'Luz', 'MONTO_BASE': 150, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 15, 'CATEGORIA': 'SERVICIOS'},
            {'ITEM': 'Agua', 'MONTO_BASE': 80, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 20, 'CATEGORIA': 'SERVICIOS'},
            {'ITEM': 'Gas', 'MONTO_BASE': 50, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 15, 'CATEGORIA': 'SERVICIOS'},
            {'ITEM': 'Internet', 'MONTO_BASE': 100, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 10, 'CATEGORIA': 'SERVICIOS'},
            {'ITEM': 'Celular Marco', 'MONTO_BASE': 50, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 5, 'CATEGORIA': 'SERVICIOS'},
            {'ITEM': 'Celular Pierina', 'MONTO_BASE': 50, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 5, 'CATEGORIA': 'SERVICIOS'},
            {'ITEM': 'Streaming (Netflix/Disney)', 'MONTO_BASE': 60, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 15, 'CATEGORIA': 'ENTRETENIMIENTO'},
            {'ITEM': 'Seguro Salud', 'MONTO_BASE': 400, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'SALUD'},
            {'ITEM': 'Colegio Hijo', 'MONTO_BASE': 800, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 5, 'CATEGORIA': 'EDUCACION'},
            {'ITEM': 'Transporte Escolar', 'MONTO_BASE': 300, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'EDUCACION'},
            {'ITEM': 'Gasolina', 'MONTO_BASE': 400, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'TRANSPORTE'},
            {'ITEM': 'Mantenimiento Auto', 'MONTO_BASE': 150, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 15, 'CATEGORIA': 'TRANSPORTE'},
            {'ITEM': 'Supermercado', 'MONTO_BASE': 1200, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'ALIMENTACION'},
            {'ITEM': 'Gasto Fijo Reserva 1', 'MONTO_BASE': 0, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'OTROS'},
            {'ITEM': 'Gasto Fijo Reserva 2', 'MONTO_BASE': 0, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'OTROS'},
            {'ITEM': 'Gasto Fijo Reserva 3', 'MONTO_BASE': 0, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'OTROS'},
        ],
        'GASTOS_VARIABLES': [
            {'ITEM': 'Restaurantes', 'CATEGORIA': 'ALIMENTACION', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Ropa y Calzado', 'CATEGORIA': 'PERSONAL', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Médico/Farmacia', 'CATEGORIA': 'SALUD', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Entretenimiento', 'CATEGORIA': 'ENTRETENIMIENTO', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Regalos', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Préstamo a NT', 'CATEGORIA': 'PRESTAMO', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Emergencias', 'CATEGORIA': 'EMERGENCIA', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Gasto Variable Reserva 1', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Gasto Variable Reserva 2', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Gasto Variable Reserva 3', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
        ],
    },
    'NEUROTEA': {
        'INGRESOS_FIJOS': [
            {'ITEM': 'Terapias Regulares', 'MONTO_BASE': 15000, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1},
            {'ITEM': 'Evaluaciones', 'MONTO_BASE': 2000, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 15},
            {'ITEM': 'Capacitaciones', 'MONTO_BASE': 1500, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1},
            {'ITEM': 'Ingreso Fijo Reserva 1', 'MONTO_BASE': 0, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1},
            {'ITEM': 'Ingreso Fijo Reserva 2', 'MONTO_BASE': 0, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1},
            {'ITEM': 'Ingreso Fijo Reserva 3', 'MONTO_BASE': 0, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1},
        ],
        'INGRESOS_VARIABLES': [
            {'ITEM': 'Talleres Especiales', 'CATEGORIA': 'SERVICIOS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Convenios Empresas', 'CATEGORIA': 'CONVENIOS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Préstamo de Familia', 'CATEGORIA': 'PRESTAMO', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Ingreso Variable Reserva 1', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Ingreso Variable Reserva 2', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Ingreso Variable Reserva 3', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
        ],
        'GASTOS_FIJOS': [
            {'ITEM': 'Alquiler Local', 'MONTO_BASE': 2500, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'LOCAL'},
            {'ITEM': 'Luz Local', 'MONTO_BASE': 300, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 15, 'CATEGORIA': 'SERVICIOS'},
            {'ITEM': 'Agua Local', 'MONTO_BASE': 100, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 20, 'CATEGORIA': 'SERVICIOS'},
            {'ITEM': 'Internet Local', 'MONTO_BASE': 150, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 10, 'CATEGORIA': 'SERVICIOS'},
            {'ITEM': 'Sueldo Marco', 'MONTO_BASE': 4000, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'PLANILLA'},
            {'ITEM': 'Sueldo Pierina', 'MONTO_BASE': 2500, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'PLANILLA'},
            {'ITEM': 'Sueldo Terapeuta 1', 'MONTO_BASE': 2000, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'PLANILLA'},
            {'ITEM': 'Sueldo Terapeuta 2', 'MONTO_BASE': 2000, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'PLANILLA'},
            {'ITEM': 'Sueldo Terapeuta 3', 'MONTO_BASE': 1800, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'PLANILLA'},
            {'ITEM': 'Contador', 'MONTO_BASE': 500, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 5, 'CATEGORIA': 'SERVICIOS_PROF'},
            {'ITEM': 'Limpieza', 'MONTO_BASE': 400, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'SERVICIOS'},
            {'ITEM': 'Material Terapéutico', 'MONTO_BASE': 500, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'MATERIALES'},
            {'ITEM': 'Marketing Digital', 'MONTO_BASE': 300, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'MARKETING'},
            {'ITEM': 'Gasto Fijo Reserva 1', 'MONTO_BASE': 0, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'OTROS'},
            {'ITEM': 'Gasto Fijo Reserva 2', 'MONTO_BASE': 0, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'OTROS'},
            {'ITEM': 'Gasto Fijo Reserva 3', 'MONTO_BASE': 0, 'FRECUENCIA': 'MENSUAL', 'DIA_VENC': 1, 'CATEGORIA': 'OTROS'},
        ],
        'GASTOS_VARIABLES': [
            {'ITEM': 'Mantenimiento Local', 'CATEGORIA': 'LOCAL', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Equipos/Mobiliario', 'CATEGORIA': 'ACTIVOS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Capacitación Personal', 'CATEGORIA': 'RRHH', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Devolución préstamo a Familia', 'CATEGORIA': 'PRESTAMO', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Impuestos/SUNAT', 'CATEGORIA': 'IMPUESTOS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Gasto Variable Reserva 1', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Gasto Variable Reserva 2', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
            {'ITEM': 'Gasto Variable Reserva 3', 'CATEGORIA': 'OTROS', 'FRECUENCIA': 'EVENTUAL'},
        ],
        'EVENTOS': [
            {'ITEM': 'Día del Niño TEA (Abril)', 'MES': 'ABR', 'PRESUPUESTO': 1500},
            {'ITEM': 'Aniversario NeuroTEA (Junio)', 'MES': 'JUN', 'PRESUPUESTO': 3000},
            {'ITEM': 'Día del Terapeuta (Sep)', 'MES': 'SEP', 'PRESUPUESTO': 1000},
            {'ITEM': 'Navidad Pacientes (Dic)', 'MES': 'DIC', 'PRESUPUESTO': 2500},
            {'ITEM': 'Capacitación Anual (Mar)', 'MES': 'MAR', 'PRESUPUESTO': 2000},
            {'ITEM': 'Fiesta Fin de Año (Dic)', 'MES': 'DIC', 'PRESUPUESTO': 2000},
            {'ITEM': 'Evento Reserva 1', 'MES': 'ENE', 'PRESUPUESTO': 0},
            {'ITEM': 'Evento Reserva 2', 'MES': 'ENE', 'PRESUPUESTO': 0},
            {'ITEM': 'Evento Reserva 3', 'MES': 'ENE', 'PRESUPUESTO': 0},
            {'ITEM': 'Evento Reserva 4', 'MES': 'ENE', 'PRESUPUESTO': 0},
            {'ITEM': 'Evento Reserva 5', 'MES': 'ENE', 'PRESUPUESTO': 0},
            {'ITEM': 'Evento Reserva 6', 'MES': 'ENE', 'PRESUPUESTO': 0},
            {'ITEM': 'Evento Reserva 7', 'MES': 'ENE', 'PRESUPUESTO': 0},
            {'ITEM': 'Evento Reserva 8', 'MES': 'ENE', 'PRESUPUESTO': 0},
            {'ITEM': 'Evento Reserva 9', 'MES': 'ENE', 'PRESUPUESTO': 0},
            {'ITEM': 'Evento Reserva 10', 'MES': 'ENE', 'PRESUPUESTO': 0},
        ],
    },
}

# ============================================================================
# ESTILOS
# ============================================================================

def crear_estilos():
    """Crear estilos reutilizables"""
    estilos = {}

    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    medium_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    estilos['thin_border'] = thin_border
    estilos['medium_border'] = medium_border

    return estilos

# ============================================================================
# FUNCIONES PARA CREAR HOJAS
# ============================================================================

def crear_hoja_tablero(wb):
    """Crear hoja TABLERO con dashboard"""
    ws = wb.create_sheet("TABLERO", 0)
    estilos = crear_estilos()

    # Configurar anchos de columna
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 3
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15

    # ========== TÍTULO PRINCIPAL ==========
    ws.merge_cells('B2:J2')
    ws['B2'] = 'CONTROL FINANCIERO 2026'
    ws['B2'].font = Font(bold=True, size=20, color='FFFFFF')
    ws['B2'].fill = PatternFill(start_color=COLORES['GRIS_HEADER'], end_color=COLORES['GRIS_HEADER'], fill_type='solid')
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 35

    # Fecha actual
    ws.merge_cells('B3:J3')
    ws['B3'] = f'Actualizado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['B3'].font = Font(italic=True, size=10, color='666666')
    ws['B3'].alignment = Alignment(horizontal='center')

    row = 5

    # ========== SECCIÓN FAMILIA (Columnas B-E) ==========
    # Header FAMILIA
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'] = 'FAMILIA'
    ws[f'B{row}'].font = Font(bold=True, size=14, color='FFFFFF')
    ws[f'B{row}'].fill = PatternFill(start_color=COLORES['VERDE_FAMILIA'], end_color=COLORES['VERDE_FAMILIA'], fill_type='solid')
    ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30

    # ========== SECCIÓN NEUROTEA (Columnas G-J) ==========
    ws.merge_cells(f'G{row}:J{row}')
    ws[f'G{row}'] = 'NEUROTEA'
    ws[f'G{row}'].font = Font(bold=True, size=14, color='FFFFFF')
    ws[f'G{row}'].fill = PatternFill(start_color=COLORES['AZUL_NEUROTEA'], end_color=COLORES['AZUL_NEUROTEA'], fill_type='solid')
    ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')

    row += 2

    # ========== SALDOS EN CUENTAS ==========
    # FAMILIA
    ws[f'B{row}'] = 'SALDOS EN CUENTAS'
    ws[f'B{row}'].font = Font(bold=True, size=11)
    ws[f'B{row}'].fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')
    ws.merge_cells(f'B{row}:E{row}')

    # NEUROTEA
    ws[f'G{row}'] = 'SALDOS EN CUENTAS'
    ws[f'G{row}'].font = Font(bold=True, size=11)
    ws[f'G{row}'].fill = PatternFill(start_color=COLORES['AZUL_CLARO'], end_color=COLORES['AZUL_CLARO'], fill_type='solid')
    ws.merge_cells(f'G{row}:J{row}')

    row += 1

    # Headers de tabla
    for col, header in [('B', 'Cuenta'), ('C', 'Esperado'), ('D', 'Real'), ('E', 'Diferencia')]:
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True, size=9)
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        ws[f'{col}{row}'].border = estilos['thin_border']

    for col, header in [('G', 'Cuenta'), ('H', 'Esperado'), ('I', 'Real'), ('J', 'Diferencia')]:
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True, size=9)
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        ws[f'{col}{row}'].border = estilos['thin_border']

    row += 1

    # Cuentas FAMILIA
    for cuenta in CONFIG['CUENTAS_FAMILIA']:
        ws[f'B{row}'] = cuenta
        ws[f'C{row}'] = 0  # Esperado
        ws[f'D{row}'] = 0  # Real
        ws[f'E{row}'] = f'=D{row}-C{row}'  # Diferencia
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].border = estilos['thin_border']
            if col in ['C', 'D', 'E']:
                ws[f'{col}{row}'].number_format = '#,##0.00'
        row += 1

    # Cuentas NEUROTEA (mismo row que familia)
    row_nt = row - len(CONFIG['CUENTAS_FAMILIA'])
    for cuenta in CONFIG['CUENTAS_NEUROTEA']:
        ws[f'G{row_nt}'] = cuenta
        ws[f'H{row_nt}'] = 0
        ws[f'I{row_nt}'] = 0
        ws[f'J{row_nt}'] = f'=I{row_nt}-H{row_nt}'
        for col in ['G', 'H', 'I', 'J']:
            ws[f'{col}{row_nt}'].border = estilos['thin_border']
            if col in ['H', 'I', 'J']:
                ws[f'{col}{row_nt}'].number_format = '#,##0.00'
        row_nt += 1

    # Total saldos
    row = max(row, row_nt) + 1
    ws[f'B{row}'] = 'TOTAL'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f'=SUM(C{row-len(CONFIG["CUENTAS_FAMILIA"])}:C{row-1})'
    ws[f'D{row}'] = f'=SUM(D{row-len(CONFIG["CUENTAS_FAMILIA"])}:D{row-1})'
    ws[f'E{row}'] = f'=D{row}-C{row}'
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = estilos['thin_border']
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')
        if col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].number_format = '#,##0.00'

    row_nt_total = row - len(CONFIG['CUENTAS_FAMILIA']) + len(CONFIG['CUENTAS_NEUROTEA'])
    ws[f'G{row}'] = 'TOTAL'
    ws[f'G{row}'].font = Font(bold=True)
    ws[f'H{row}'] = f'=SUM(H{row-len(CONFIG["CUENTAS_NEUROTEA"])-1}:H{row-1})'
    ws[f'I{row}'] = f'=SUM(I{row-len(CONFIG["CUENTAS_NEUROTEA"])-1}:I{row-1})'
    ws[f'J{row}'] = f'=I{row}-H{row}'
    for col in ['G', 'H', 'I', 'J']:
        ws[f'{col}{row}'].border = estilos['thin_border']
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORES['AZUL_CLARO'], end_color=COLORES['AZUL_CLARO'], fill_type='solid')
        if col in ['H', 'I', 'J']:
            ws[f'{col}{row}'].number_format = '#,##0.00'

    row += 2

    # ========== PRESUPUESTO VS REAL ==========
    # FAMILIA
    ws[f'B{row}'] = 'PRESUPUESTO VS REAL'
    ws[f'B{row}'].font = Font(bold=True, size=11)
    ws[f'B{row}'].fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')
    ws.merge_cells(f'B{row}:E{row}')

    # NEUROTEA
    ws[f'G{row}'] = 'PRESUPUESTO VS REAL'
    ws[f'G{row}'].font = Font(bold=True, size=11)
    ws[f'G{row}'].fill = PatternFill(start_color=COLORES['AZUL_CLARO'], end_color=COLORES['AZUL_CLARO'], fill_type='solid')
    ws.merge_cells(f'G{row}:J{row}')

    row += 1

    # Headers
    for col, header in [('B', 'Concepto'), ('C', 'Presupuesto'), ('D', 'Real'), ('E', '% Ejec')]:
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True, size=9)
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        ws[f'{col}{row}'].border = estilos['thin_border']

    for col, header in [('G', 'Concepto'), ('H', 'Presupuesto'), ('I', 'Real'), ('J', '% Ejec')]:
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True, size=9)
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        ws[f'{col}{row}'].border = estilos['thin_border']

    row += 1

    # Datos presupuesto FAMILIA
    conceptos_fam = ['Ingresos Fijos', 'Ingresos Variables', 'Gastos Fijos', 'Gastos Variables']
    for concepto in conceptos_fam:
        ws[f'B{row}'] = concepto
        ws[f'C{row}'] = 0
        ws[f'D{row}'] = 0
        ws[f'E{row}'] = f'=IF(C{row}=0,0,D{row}/C{row})'
        ws[f'E{row}'].number_format = '0%'
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].border = estilos['thin_border']
            if col in ['C', 'D']:
                ws[f'{col}{row}'].number_format = '#,##0.00'
        row += 1

    # Datos presupuesto NEUROTEA
    row_nt = row - 4
    conceptos_nt = ['Ingresos Fijos', 'Ingresos Variables', 'Gastos Fijos', 'Gastos Variables', 'Eventos']
    for concepto in conceptos_nt:
        ws[f'G{row_nt}'] = concepto
        ws[f'H{row_nt}'] = 0
        ws[f'I{row_nt}'] = 0
        ws[f'J{row_nt}'] = f'=IF(H{row_nt}=0,0,I{row_nt}/H{row_nt})'
        ws[f'J{row_nt}'].number_format = '0%'
        for col in ['G', 'H', 'I', 'J']:
            ws[f'{col}{row_nt}'].border = estilos['thin_border']
            if col in ['H', 'I']:
                ws[f'{col}{row_nt}'].number_format = '#,##0.00'
        row_nt += 1

    row = max(row, row_nt) + 1

    # ========== FLUJO DEL MES ==========
    ws[f'B{row}'] = 'FLUJO DEL MES'
    ws[f'B{row}'].font = Font(bold=True, size=11)
    ws[f'B{row}'].fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')
    ws.merge_cells(f'B{row}:E{row}')

    ws[f'G{row}'] = 'FLUJO DEL MES'
    ws[f'G{row}'].font = Font(bold=True, size=11)
    ws[f'G{row}'].fill = PatternFill(start_color=COLORES['AZUL_CLARO'], end_color=COLORES['AZUL_CLARO'], fill_type='solid')
    ws.merge_cells(f'G{row}:J{row}')

    row += 1

    flujo_items = ['Total Ingresos', 'Total Egresos', 'Flujo Neto']
    for item in flujo_items:
        ws[f'B{row}'] = item
        ws[f'C{row}'] = 0
        ws[f'G{row}'] = item
        ws[f'H{row}'] = 0
        for col in ['B', 'C']:
            ws[f'{col}{row}'].border = estilos['thin_border']
        for col in ['G', 'H']:
            ws[f'{col}{row}'].border = estilos['thin_border']
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'H{row}'].number_format = '#,##0.00'
        if item == 'Flujo Neto':
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'G{row}'].font = Font(bold=True)
            ws[f'C{row}'].fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')
            ws[f'H{row}'].fill = PatternFill(start_color=COLORES['AZUL_CLARO'], end_color=COLORES['AZUL_CLARO'], fill_type='solid')
        row += 1

    row += 1

    # ========== LIQUIDEZ 3 SEMANAS ==========
    ws[f'B{row}'] = 'LIQUIDEZ PRÓXIMAS 3 SEMANAS'
    ws[f'B{row}'].font = Font(bold=True, size=11)
    ws[f'B{row}'].fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')
    ws.merge_cells(f'B{row}:E{row}')

    ws[f'G{row}'] = 'LIQUIDEZ PRÓXIMAS 3 SEMANAS'
    ws[f'G{row}'].font = Font(bold=True, size=11)
    ws[f'G{row}'].fill = PatternFill(start_color=COLORES['AZUL_CLARO'], end_color=COLORES['AZUL_CLARO'], fill_type='solid')
    ws.merge_cells(f'G{row}:J{row}')

    row += 1

    # Headers liquidez
    for col, header in [('B', 'Semana'), ('C', 'Ingresos'), ('D', 'Egresos'), ('E', 'Saldo')]:
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True, size=9)
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        ws[f'{col}{row}'].border = estilos['thin_border']

    for col, header in [('G', 'Semana'), ('H', 'Ingresos'), ('I', 'Egresos'), ('J', 'Saldo')]:
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True, size=9)
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        ws[f'{col}{row}'].border = estilos['thin_border']

    row += 1

    for semana in ['Semana 1', 'Semana 2', 'Semana 3']:
        ws[f'B{row}'] = semana
        ws[f'C{row}'] = 0
        ws[f'D{row}'] = 0
        ws[f'E{row}'] = f'=C{row}-D{row}'
        ws[f'G{row}'] = semana
        ws[f'H{row}'] = 0
        ws[f'I{row}'] = 0
        ws[f'J{row}'] = f'=H{row}-I{row}'
        for col in ['B', 'C', 'D', 'E', 'G', 'H', 'I', 'J']:
            ws[f'{col}{row}'].border = estilos['thin_border']
            if col in ['C', 'D', 'E', 'H', 'I', 'J']:
                ws[f'{col}{row}'].number_format = '#,##0.00'
        row += 1

    row += 2

    # ========== BALANCE CRUZADO ==========
    ws.merge_cells(f'B{row}:J{row}')
    ws[f'B{row}'] = 'BALANCE CRUZADO NT <-> FAMILIA'
    ws[f'B{row}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws[f'B{row}'].fill = PatternFill(start_color=COLORES['GRIS_HEADER'], end_color=COLORES['GRIS_HEADER'], fill_type='solid')
    ws[f'B{row}'].alignment = Alignment(horizontal='center')

    row += 1

    # Headers balance cruzado
    for col, header in [('B', 'Concepto'), ('C', 'NT debe a FAM'), ('D', 'FAM debe a NT'), ('E', 'Saldo Neto')]:
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True, size=9)
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        ws[f'{col}{row}'].border = estilos['thin_border']

    row += 1

    ws[f'B{row}'] = 'Préstamos Activos'
    ws[f'C{row}'] = 0
    ws[f'D{row}'] = 0
    ws[f'E{row}'] = '=C' + str(row) + '-D' + str(row)
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = estilos['thin_border']
        if col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].number_format = '#,##0.00'

    row += 1
    ws[f'B{row}'] = 'BALANCE'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'E{row}'] = '=E' + str(row-1)
    ws[f'E{row}'].font = Font(bold=True, size=12)
    ws[f'E{row}'].number_format = '#,##0.00'
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].border = estilos['thin_border']
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORES['AMARILLO_CLARO'], end_color=COLORES['AMARILLO_CLARO'], fill_type='solid')

    return ws


def crear_hoja_movimiento(wb):
    """Crear hoja MOVIMIENTO para registro de transacciones"""
    ws = wb.create_sheet("MOVIMIENTO")
    estilos = crear_estilos()

    # Configurar anchos
    anchos = {'A': 12, 'B': 12, 'C': 15, 'D': 12, 'E': 30, 'F': 15, 'G': 15,
              'H': 12, 'I': 15, 'J': 12, 'K': 25, 'L': 12}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    # Headers
    headers = ['FECHA', 'ENTIDAD', 'TIPO', 'CUENTA', 'ITEM', 'CATEGORIA',
               'MONTO', 'DIA_VENC', 'ESTADO', 'MES', 'DESCRIPCION', 'ID']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORES['GRIS_HEADER'], end_color=COLORES['GRIS_HEADER'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = estilos['thin_border']

    # Congelar primera fila
    ws.freeze_panes = 'A2'

    # Crear validaciones de datos
    # Entidad
    dv_entidad = DataValidation(type="list", formula1='"FAMILIA,NEUROTEA"', allow_blank=False)
    dv_entidad.error = "Seleccione una entidad válida"
    dv_entidad.errorTitle = "Entidad inválida"
    ws.add_data_validation(dv_entidad)
    dv_entidad.add('B2:B1000')

    # Tipo
    dv_tipo = DataValidation(type="list", formula1='"INGRESO,EGRESO,TRANSFERENCIA"', allow_blank=False)
    ws.add_data_validation(dv_tipo)
    dv_tipo.add('C2:C1000')

    # Estado
    dv_estado = DataValidation(type="list", formula1='"PENDIENTE,PAGADO,ANULADO"', allow_blank=False)
    ws.add_data_validation(dv_estado)
    dv_estado.add('I2:I1000')

    # Mes
    dv_mes = DataValidation(type="list", formula1='"ENE,FEB,MAR,ABR,MAY,JUN,JUL,AGO,SEP,OCT,NOV,DIC"', allow_blank=False)
    ws.add_data_validation(dv_mes)
    dv_mes.add('J2:J1000')

    # Formato de fecha y número para las primeras 100 filas
    for row in range(2, 102):
        ws.cell(row=row, column=1).number_format = 'DD/MM/YYYY'
        ws.cell(row=row, column=7).number_format = '#,##0.00'
        ws.cell(row=row, column=8).number_format = '0'

    return ws


def crear_hoja_presupuesto(wb, entidad):
    """Crear hoja de presupuesto para una entidad"""
    nombre_hoja = f"PPTO_{entidad[:3]}"
    ws = wb.create_sheet(nombre_hoja)
    estilos = crear_estilos()

    color_header = COLORES['VERDE_FAMILIA'] if entidad == 'FAMILIA' else COLORES['AZUL_NEUROTEA']
    color_claro = COLORES['VERDE_CLARO'] if entidad == 'FAMILIA' else COLORES['AZUL_CLARO']

    # Configurar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    for i, mes in enumerate(CONFIG['MESES']):
        ws.column_dimensions[get_column_letter(6 + i)].width = 10

    row = 1

    # Título
    ws.merge_cells(f'A{row}:Q{row}')
    ws[f'A{row}'] = f'PRESUPUESTO {entidad} 2026'
    ws[f'A{row}'].font = Font(bold=True, size=14, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 25

    row += 2

    datos = DATOS[entidad]

    # ========== INGRESOS FIJOS ==========
    ws[f'A{row}'] = 'INGRESOS FIJOS'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color=color_claro, end_color=color_claro, fill_type='solid')
    ws.merge_cells(f'A{row}:Q{row}')
    row += 1

    # Headers
    headers = ['ITEM', 'MONTO_BASE', 'FRECUENCIA', 'DIA_VENC', 'ANUAL'] + CONFIG['MESES']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = estilos['thin_border']
    row += 1

    start_row_if = row
    for item in datos['INGRESOS_FIJOS']:
        ws.cell(row=row, column=1, value=item['ITEM']).border = estilos['thin_border']
        ws.cell(row=row, column=2, value=item['MONTO_BASE']).border = estilos['thin_border']
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=item['FRECUENCIA']).border = estilos['thin_border']
        ws.cell(row=row, column=4, value=item['DIA_VENC']).border = estilos['thin_border']
        ws.cell(row=row, column=5, value=f'=B{row}*12').border = estilos['thin_border']
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        # Meses
        for col in range(6, 18):
            ws.cell(row=row, column=col, value=f'=B{row}').border = estilos['thin_border']
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        row += 1

    # Total Ingresos Fijos
    ws.cell(row=row, column=1, value='TOTAL INGRESOS FIJOS').font = Font(bold=True)
    for col in range(2, 18):
        if col == 2:
            ws.cell(row=row, column=col, value=f'=SUM(B{start_row_if}:B{row-1})')
        elif col == 5:
            ws.cell(row=row, column=col, value=f'=SUM(E{start_row_if}:E{row-1})')
        elif col >= 6:
            col_letter = get_column_letter(col)
            ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_row_if}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = estilos['thin_border']
        ws.cell(row=row, column=col).number_format = '#,##0.00'
        ws.cell(row=row, column=col).fill = PatternFill(start_color=color_claro, end_color=color_claro, fill_type='solid')
    row += 2

    # ========== INGRESOS VARIABLES ==========
    ws[f'A{row}'] = 'INGRESOS VARIABLES'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color=color_claro, end_color=color_claro, fill_type='solid')
    ws.merge_cells(f'A{row}:Q{row}')
    row += 1

    headers_var = ['ITEM', 'CATEGORIA', 'FRECUENCIA', '', 'ANUAL'] + CONFIG['MESES']
    for col_num, header in enumerate(headers_var, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = estilos['thin_border']
    row += 1

    start_row_iv = row
    for item in datos['INGRESOS_VARIABLES']:
        ws.cell(row=row, column=1, value=item['ITEM']).border = estilos['thin_border']
        ws.cell(row=row, column=2, value=item['CATEGORIA']).border = estilos['thin_border']
        ws.cell(row=row, column=3, value=item['FRECUENCIA']).border = estilos['thin_border']
        ws.cell(row=row, column=5, value=f'=SUM(F{row}:Q{row})').border = estilos['thin_border']
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        for col in range(6, 18):
            ws.cell(row=row, column=col, value=0).border = estilos['thin_border']
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        row += 1

    # Total Ingresos Variables
    ws.cell(row=row, column=1, value='TOTAL INGRESOS VARIABLES').font = Font(bold=True)
    for col in range(5, 18):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_row_iv}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = estilos['thin_border']
        ws.cell(row=row, column=col).number_format = '#,##0.00'
        ws.cell(row=row, column=col).fill = PatternFill(start_color=color_claro, end_color=color_claro, fill_type='solid')
    row += 2

    # ========== GASTOS FIJOS ==========
    ws[f'A{row}'] = 'GASTOS FIJOS'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['ROJO_CLARO'], end_color=COLORES['ROJO_CLARO'], fill_type='solid')
    ws.merge_cells(f'A{row}:Q{row}')
    row += 1

    headers_gf = ['ITEM', 'MONTO_BASE', 'FRECUENCIA', 'DIA_VENC', 'ANUAL'] + CONFIG['MESES']
    for col_num, header in enumerate(headers_gf, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = estilos['thin_border']
    row += 1

    start_row_gf = row
    for item in datos['GASTOS_FIJOS']:
        ws.cell(row=row, column=1, value=item['ITEM']).border = estilos['thin_border']
        ws.cell(row=row, column=2, value=item['MONTO_BASE']).border = estilos['thin_border']
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=item['FRECUENCIA']).border = estilos['thin_border']
        ws.cell(row=row, column=4, value=item['DIA_VENC']).border = estilos['thin_border']
        ws.cell(row=row, column=5, value=f'=B{row}*12').border = estilos['thin_border']
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        for col in range(6, 18):
            ws.cell(row=row, column=col, value=f'=B{row}').border = estilos['thin_border']
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        row += 1

    # Total Gastos Fijos
    ws.cell(row=row, column=1, value='TOTAL GASTOS FIJOS').font = Font(bold=True)
    for col in range(2, 18):
        if col == 2:
            ws.cell(row=row, column=col, value=f'=SUM(B{start_row_gf}:B{row-1})')
        elif col == 5:
            ws.cell(row=row, column=col, value=f'=SUM(E{start_row_gf}:E{row-1})')
        elif col >= 6:
            col_letter = get_column_letter(col)
            ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_row_gf}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = estilos['thin_border']
        ws.cell(row=row, column=col).number_format = '#,##0.00'
        ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORES['ROJO_CLARO'], end_color=COLORES['ROJO_CLARO'], fill_type='solid')
    row += 2

    # ========== GASTOS VARIABLES ==========
    ws[f'A{row}'] = 'GASTOS VARIABLES'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['ROJO_CLARO'], end_color=COLORES['ROJO_CLARO'], fill_type='solid')
    ws.merge_cells(f'A{row}:Q{row}')
    row += 1

    for col_num, header in enumerate(headers_var, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = estilos['thin_border']
    row += 1

    start_row_gv = row
    for item in datos['GASTOS_VARIABLES']:
        ws.cell(row=row, column=1, value=item['ITEM']).border = estilos['thin_border']
        ws.cell(row=row, column=2, value=item['CATEGORIA']).border = estilos['thin_border']
        ws.cell(row=row, column=3, value=item['FRECUENCIA']).border = estilos['thin_border']
        ws.cell(row=row, column=5, value=f'=SUM(F{row}:Q{row})').border = estilos['thin_border']
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        for col in range(6, 18):
            ws.cell(row=row, column=col, value=0).border = estilos['thin_border']
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        row += 1

    # Total Gastos Variables
    ws.cell(row=row, column=1, value='TOTAL GASTOS VARIABLES').font = Font(bold=True)
    for col in range(5, 18):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_row_gv}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = estilos['thin_border']
        ws.cell(row=row, column=col).number_format = '#,##0.00'
        ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORES['ROJO_CLARO'], end_color=COLORES['ROJO_CLARO'], fill_type='solid')
    row += 2

    # ========== EVENTOS (Solo NEUROTEA) ==========
    if entidad == 'NEUROTEA' and 'EVENTOS' in datos:
        ws[f'A{row}'] = 'EVENTOS'
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = PatternFill(start_color=COLORES['AMARILLO_CLARO'], end_color=COLORES['AMARILLO_CLARO'], fill_type='solid')
        ws.merge_cells(f'A{row}:Q{row}')
        row += 1

        headers_ev = ['EVENTO', 'MES', 'PRESUPUESTO', '', 'ANUAL'] + CONFIG['MESES']
        for col_num, header in enumerate(headers_ev, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
            cell.border = estilos['thin_border']
        row += 1

        start_row_ev = row
        for evento in datos['EVENTOS']:
            ws.cell(row=row, column=1, value=evento['ITEM']).border = estilos['thin_border']
            ws.cell(row=row, column=2, value=evento['MES']).border = estilos['thin_border']
            ws.cell(row=row, column=3, value=evento['PRESUPUESTO']).border = estilos['thin_border']
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=f'=C{row}').border = estilos['thin_border']
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            # Poner el monto en el mes correspondiente
            mes_idx = CONFIG['MESES'].index(evento['MES']) if evento['MES'] in CONFIG['MESES'] else 0
            for col in range(6, 18):
                if col - 6 == mes_idx:
                    ws.cell(row=row, column=col, value=f'=C{row}').border = estilos['thin_border']
                else:
                    ws.cell(row=row, column=col, value=0).border = estilos['thin_border']
                ws.cell(row=row, column=col).number_format = '#,##0.00'
            row += 1

        # Total Eventos
        ws.cell(row=row, column=1, value='TOTAL EVENTOS').font = Font(bold=True)
        for col in range(3, 18):
            if col in [3, 5]:
                col_letter = get_column_letter(col)
                ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_row_ev}:{col_letter}{row-1})')
            elif col >= 6:
                col_letter = get_column_letter(col)
                ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_row_ev}:{col_letter}{row-1})')
            ws.cell(row=row, column=col).border = estilos['thin_border']
            ws.cell(row=row, column=col).number_format = '#,##0.00'
            ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORES['AMARILLO_CLARO'], end_color=COLORES['AMARILLO_CLARO'], fill_type='solid')

    return ws


def crear_hoja_config(wb):
    """Crear hoja CONFIG con listas de validación"""
    ws = wb.create_sheet("CONFIG")
    estilos = crear_estilos()

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = 'CONFIGURACIÓN DEL SISTEMA'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color=COLORES['GRIS_HEADER'], end_color=COLORES['GRIS_HEADER'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Crear listas
    col = 1
    for nombre, valores in CONFIG.items():
        ws.cell(row=3, column=col, value=nombre).font = Font(bold=True)
        ws.cell(row=3, column=col).fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        for i, valor in enumerate(valores):
            ws.cell(row=4+i, column=col, value=valor)
        col += 1

    # Ajustar anchos
    for i in range(1, col):
        ws.column_dimensions[get_column_letter(i)].width = 18

    return ws


def crear_hoja_cuentas(wb):
    """Crear hoja CUENTAS para saldos bancarios"""
    ws = wb.create_sheet("CUENTAS")
    estilos = crear_estilos()

    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = 'SALDOS DE CUENTAS BANCARIAS'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color=COLORES['GRIS_HEADER'], end_color=COLORES['GRIS_HEADER'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['ENTIDAD', 'CUENTA', 'SALDO_INICIAL', 'SALDO_REAL', 'FECHA_ACT', 'NOTAS']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = estilos['thin_border']

    # Cuentas FAMILIA
    row = 4
    for cuenta in CONFIG['CUENTAS_FAMILIA']:
        ws.cell(row=row, column=1, value='FAMILIA').border = estilos['thin_border']
        ws.cell(row=row, column=2, value=cuenta).border = estilos['thin_border']
        ws.cell(row=row, column=3, value=0).border = estilos['thin_border']
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=0).border = estilos['thin_border']
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=datetime.now().strftime('%d/%m/%Y')).border = estilos['thin_border']
        ws.cell(row=row, column=6, value='').border = estilos['thin_border']
        row += 1

    # Cuentas NEUROTEA
    for cuenta in CONFIG['CUENTAS_NEUROTEA']:
        ws.cell(row=row, column=1, value='NEUROTEA').border = estilos['thin_border']
        ws.cell(row=row, column=2, value=cuenta).border = estilos['thin_border']
        ws.cell(row=row, column=3, value=0).border = estilos['thin_border']
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=0).border = estilos['thin_border']
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=datetime.now().strftime('%d/%m/%Y')).border = estilos['thin_border']
        ws.cell(row=row, column=6, value='').border = estilos['thin_border']
        row += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 30

    return ws


def crear_hoja_items(wb):
    """Crear hoja ITEMS con todos los items del presupuesto"""
    ws = wb.create_sheet("ITEMS")
    estilos = crear_estilos()

    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = 'CATÁLOGO DE ITEMS'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color=COLORES['GRIS_HEADER'], end_color=COLORES['GRIS_HEADER'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['ENTIDAD', 'TIPO', 'CLASIFICACION', 'ITEM', 'CATEGORIA', 'FRECUENCIA']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = estilos['thin_border']

    row = 4

    for entidad in ['FAMILIA', 'NEUROTEA']:
        datos = DATOS[entidad]
        color = COLORES['VERDE_CLARO'] if entidad == 'FAMILIA' else COLORES['AZUL_CLARO']

        # Ingresos Fijos
        for item in datos['INGRESOS_FIJOS']:
            ws.cell(row=row, column=1, value=entidad).border = estilos['thin_border']
            ws.cell(row=row, column=2, value='INGRESO').border = estilos['thin_border']
            ws.cell(row=row, column=3, value='FIJO').border = estilos['thin_border']
            ws.cell(row=row, column=4, value=item['ITEM']).border = estilos['thin_border']
            ws.cell(row=row, column=5, value='INGRESO_FIJO').border = estilos['thin_border']
            ws.cell(row=row, column=6, value=item['FRECUENCIA']).border = estilos['thin_border']
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            row += 1

        # Ingresos Variables
        for item in datos['INGRESOS_VARIABLES']:
            ws.cell(row=row, column=1, value=entidad).border = estilos['thin_border']
            ws.cell(row=row, column=2, value='INGRESO').border = estilos['thin_border']
            ws.cell(row=row, column=3, value='VARIABLE').border = estilos['thin_border']
            ws.cell(row=row, column=4, value=item['ITEM']).border = estilos['thin_border']
            ws.cell(row=row, column=5, value=item['CATEGORIA']).border = estilos['thin_border']
            ws.cell(row=row, column=6, value=item['FRECUENCIA']).border = estilos['thin_border']
            row += 1

        # Gastos Fijos
        for item in datos['GASTOS_FIJOS']:
            ws.cell(row=row, column=1, value=entidad).border = estilos['thin_border']
            ws.cell(row=row, column=2, value='EGRESO').border = estilos['thin_border']
            ws.cell(row=row, column=3, value='FIJO').border = estilos['thin_border']
            ws.cell(row=row, column=4, value=item['ITEM']).border = estilos['thin_border']
            ws.cell(row=row, column=5, value=item.get('CATEGORIA', 'OTROS')).border = estilos['thin_border']
            ws.cell(row=row, column=6, value=item['FRECUENCIA']).border = estilos['thin_border']
            row += 1

        # Gastos Variables
        for item in datos['GASTOS_VARIABLES']:
            ws.cell(row=row, column=1, value=entidad).border = estilos['thin_border']
            ws.cell(row=row, column=2, value='EGRESO').border = estilos['thin_border']
            ws.cell(row=row, column=3, value='VARIABLE').border = estilos['thin_border']
            ws.cell(row=row, column=4, value=item['ITEM']).border = estilos['thin_border']
            ws.cell(row=row, column=5, value=item['CATEGORIA']).border = estilos['thin_border']
            ws.cell(row=row, column=6, value=item['FRECUENCIA']).border = estilos['thin_border']
            row += 1

        # Eventos (solo NEUROTEA)
        if entidad == 'NEUROTEA' and 'EVENTOS' in datos:
            for evento in datos['EVENTOS']:
                ws.cell(row=row, column=1, value=entidad).border = estilos['thin_border']
                ws.cell(row=row, column=2, value='EGRESO').border = estilos['thin_border']
                ws.cell(row=row, column=3, value='EVENTO').border = estilos['thin_border']
                ws.cell(row=row, column=4, value=evento['ITEM']).border = estilos['thin_border']
                ws.cell(row=row, column=5, value='EVENTO').border = estilos['thin_border']
                ws.cell(row=row, column=6, value='EVENTUAL').border = estilos['thin_border']
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORES['AMARILLO_CLARO'], end_color=COLORES['AMARILLO_CLARO'], fill_type='solid')
                row += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12

    return ws


def crear_hoja_instrucciones(wb):
    """Crear hoja de instrucciones"""
    ws = wb.create_sheet("INSTRUCCIONES")

    ws.column_dimensions['A'].width = 100

    instrucciones = [
        "═══════════════════════════════════════════════════════════════════════════════",
        "                    CONTROL FINANCIERO 2026 - INSTRUCCIONES",
        "═══════════════════════════════════════════════════════════════════════════════",
        "",
        "ESTRUCTURA DE HOJAS:",
        "─────────────────────",
        "• TABLERO      → Dashboard visual con resumen de ambas entidades",
        "• MOVIMIENTO   → Registro de todas las transacciones",
        "• PPTO_FAM     → Presupuesto detallado de FAMILIA",
        "• PPTO_NEU     → Presupuesto detallado de NEUROTEA",
        "• CUENTAS      → Saldos de cuentas bancarias",
        "• ITEMS        → Catálogo completo de items",
        "• CONFIG       → Listas de configuración",
        "",
        "CÓMO USAR:",
        "───────────",
        "1. Actualizar saldos reales en hoja CUENTAS",
        "2. Registrar movimientos en hoja MOVIMIENTO",
        "3. Revisar presupuesto en PPTO_FAM y PPTO_NEU",
        "4. Consultar TABLERO para ver resumen",
        "",
        "VALIDACIONES:",
        "──────────────",
        "• ENTIDAD: Solo FAMILIA o NEUROTEA",
        "• TIPO: INGRESO, EGRESO, TRANSFERENCIA",
        "• ESTADO: PENDIENTE, PAGADO, ANULADO",
        "• MES: ENE, FEB, MAR, ABR, MAY, JUN, JUL, AGO, SEP, OCT, NOV, DIC",
        "",
        "BALANCE CRUZADO:",
        "─────────────────",
        "• Los sueldos de Marco y Pierina son EGRESO de NEUROTEA",
        "• Los mismos sueldos son INGRESO de FAMILIA",
        "• El préstamo de FAM a NT se registra como:",
        "  - EGRESO en FAMILIA (Préstamo a NT)",
        "  - INGRESO en NEUROTEA (Préstamo de Familia)",
        "",
        "RESERVAS:",
        "──────────",
        "• Cada categoría tiene 3 items de reserva para futuras necesidades",
        "• Los eventos de NEUROTEA tienen 10 reservas adicionales",
        "• Los items reserva tienen MONTO_BASE = 0, activar cuando se necesiten",
        "",
        "═══════════════════════════════════════════════════════════════════════════════",
        "                         Versión 3.0 - Diciembre 2024",
        "═══════════════════════════════════════════════════════════════════════════════",
    ]

    for i, linea in enumerate(instrucciones, 1):
        ws.cell(row=i, column=1, value=linea)
        if i <= 3 or i >= len(instrucciones) - 1:
            ws.cell(row=i, column=1).font = Font(bold=True, size=12)
        elif linea.endswith(':'):
            ws.cell(row=i, column=1).font = Font(bold=True, size=11)

    return ws


# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================

def crear_excel():
    """Crear el archivo Excel completo"""
    print("Creando Control Financiero 2026...")

    wb = Workbook()

    # Eliminar hoja por defecto
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Crear todas las hojas
    print("  → Creando TABLERO...")
    crear_hoja_tablero(wb)

    print("  → Creando MOVIMIENTO...")
    crear_hoja_movimiento(wb)

    print("  → Creando PPTO_FAM...")
    crear_hoja_presupuesto(wb, 'FAMILIA')

    print("  → Creando PPTO_NEU...")
    crear_hoja_presupuesto(wb, 'NEUROTEA')

    print("  → Creando CUENTAS...")
    crear_hoja_cuentas(wb)

    print("  → Creando ITEMS...")
    crear_hoja_items(wb)

    print("  → Creando CONFIG...")
    crear_hoja_config(wb)

    print("  → Creando INSTRUCCIONES...")
    crear_hoja_instrucciones(wb)

    # Guardar
    filename = "Control_Financiero_2026_V3.xlsx"
    wb.save(filename)
    print(f"\n✅ Archivo creado: {filename}")

    return filename


def verificar_archivo(filename):
    """Verificar que el archivo se creó correctamente"""
    print(f"\n📋 Verificando {filename}...")

    wb = openpyxl.load_workbook(filename)

    print(f"  Hojas encontradas: {len(wb.sheetnames)}")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"    • {sheet}: {ws.max_row} filas x {ws.max_column} columnas")

    # Verificar validaciones en MOVIMIENTO
    ws_mov = wb['MOVIMIENTO']
    print(f"  Validaciones en MOVIMIENTO: {len(ws_mov.data_validations.dataValidation)}")

    # Verificar items en ITEMS
    ws_items = wb['ITEMS']
    items_count = ws_items.max_row - 3  # Menos headers
    print(f"  Items en catálogo: {items_count}")

    # Verificar eventos en PPTO_NEU
    ws_neu = wb['PPTO_NEU']
    print(f"  PPTO_NEU filas: {ws_neu.max_row}")

    wb.close()
    print("\n✅ Verificación completada")


if __name__ == "__main__":
    filename = crear_excel()
    verificar_archivo(filename)
