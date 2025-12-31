#!/usr/bin/env python3
"""
Control Financiero 2026 - Generador Excel V4
Réplica exacta del V9 del usuario
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from datetime import datetime
from copy import copy

# ============================================================================
# COLORES (Exactos del diseño)
# ============================================================================
COLORES = {
    'VERDE_HEADER': '059669',      # Header FAMILIA
    'VERDE_CLARO': 'D1FAE5',       # Fondo verde claro
    'AZUL_HEADER': '1E40AF',       # Header NEUROTEA
    'AZUL_CLARO': 'DBEAFE',        # Fondo azul claro
    'GRIS_HEADER': '1F2937',       # Header principal oscuro
    'GRIS_CLARO': 'F3F4F6',        # Fondo gris
    'BLANCO': 'FFFFFF',
    'ROJO': 'DC2626',
    'ROJO_CLARO': 'FEE2E2',
    'AMARILLO': 'F59E0B',
    'MORADO': '7C3AED',            # Balance cruzado
    'MORADO_CLARO': 'EDE9FE',
}

# ============================================================================
# DATOS MAESTROS (Exactos del V9)
# ============================================================================

CONFIG = {
    'ENTIDADES': ['Familia', 'NeuroTEA'],
    'TIPO_MOV': ['Ingreso', 'Egreso', 'Prestamo NT-Fam', 'Devol Fam-NT'],
    'FRECUENCIA': ['Fijo/Mensual', 'Variable/Mensual', 'Fijo/Anual', 'Variable/Anual', 'Variable'],
    'ESTADO': ['Pendiente', 'Pagado', 'Cancelado'],
    'CUENTAS': ['Atlas (NT)', 'ITAU Marco', 'ITAU Clara', 'UENO Clara', 'Coop Univ Marco',
                'Coop Univ Clara', 'Caja NT', 'Efectivo'],
    'MESES': ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
              'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'],
    'MESES_CORTO': ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC'],
    'META': {
        '% Ganancia Meta': 0.07,
        '% Utilidad': 0.0233,
        '% Fondo Emerg': 0.0233,
        '% Fondo Inv': 0.0233,
        '% Gastos Max': 0.93,
    },
    'CATEGORIAS_FAMILIA': ['Gastos Fijos', 'Cuotas y Prestamos', 'Obligaciones', 'Suscripciones', 'Variables', 'Ahorro'],
    'CATEGORIAS_NT': ['Clinica', 'Sueldos y Honorarios', 'Telefonia', 'Obligaciones', 'Eventos', 'Variables', 'Ganancia'],
}

# Ingresos Familia
INGRESOS_FAMILIA = [
    {'CONCEPTO': 'Salario Marco', 'TIPO': 'Fijo/Mensual', 'MONTO': 8500000},
    {'CONCEPTO': 'Salario Marco NeuroTEA', 'TIPO': 'Fijo/Mensual', 'MONTO': 5000000},
    {'CONCEPTO': 'Vacaciones Marco', 'TIPO': 'Variable/Anual', 'MONTO': 0},
    {'CONCEPTO': 'Aguinaldo Marco', 'TIPO': 'Variable/Anual', 'MONTO': 0},
    {'CONCEPTO': 'Viatico Marco', 'TIPO': 'Variable', 'MONTO': 500000},
    {'CONCEPTO': 'Animador Biblico', 'TIPO': 'Variable/Mensual', 'MONTO': 200000},
    {'CONCEPTO': 'Honorarios Clara NT', 'TIPO': 'Variable', 'MONTO': 3000000},
    {'CONCEPTO': 'Reserva', 'TIPO': '', 'MONTO': 0},
]

# Gastos Fijos Familia
GASTOS_FIJOS_FAMILIA = [
    {'CONCEPTO': 'Salario Lili Domestico', 'CATEGORIA': 'Gastos Fijos', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 30, 'BASE': 1800000},
    {'CONCEPTO': 'Salario Laura Domestico', 'CATEGORIA': 'Gastos Fijos', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 30, 'BASE': 1200000},
    {'CONCEPTO': 'Escuela Fabian y Brenda', 'CATEGORIA': 'Gastos Fijos', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 10, 'BASE': 1600000},
    {'CONCEPTO': 'Robotica Ninos', 'CATEGORIA': 'Gastos Fijos', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 11, 'BASE': 400000},
    {'CONCEPTO': 'ANDE Casa', 'CATEGORIA': 'Gastos Fijos', 'FRECUENCIA': 'Variable/Mensual', 'DIA': 15, 'BASE': 400000},
    {'CONCEPTO': 'Expensa Casa', 'CATEGORIA': 'Gastos Fijos', 'FRECUENCIA': 'Variable/Mensual', 'DIA': 10, 'BASE': 300000},
    {'CONCEPTO': 'Na Luisa', 'CATEGORIA': 'Gastos Fijos', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 5, 'BASE': 1200000},
    {'CONCEPTO': 'Contadora Marco', 'CATEGORIA': 'Gastos Fijos', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 15, 'BASE': 350000},
]

# Cuotas y Préstamos Familia
CUOTAS_FAMILIA = [
    {'CONCEPTO': 'Cajubi Marco', 'CATEGORIA': 'Cuotas', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 10, 'BASE': 200000},
    {'CONCEPTO': 'Mutual Marco', 'CATEGORIA': 'Cuotas', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 10, 'BASE': 150000},
    {'CONCEPTO': 'Cuota ITAU', 'CATEGORIA': 'Cuotas', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 15, 'BASE': 391000},
    {'CONCEPTO': 'Auto Laura Cuota', 'CATEGORIA': 'Cuotas', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 20, 'BASE': 2000000},
    {'CONCEPTO': 'Coop. Universitaria Clara', 'CATEGORIA': 'Cuotas', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 15, 'BASE': 800000},
    {'CONCEPTO': 'Coomecipar Clara', 'CATEGORIA': 'Cuotas', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 15, 'BASE': 500000},
    {'CONCEPTO': 'Solar Prestamo 1', 'CATEGORIA': 'Cuotas', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 20, 'BASE': 850000},
    {'CONCEPTO': 'Solar Prestamo 2', 'CATEGORIA': 'Cuotas', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 20, 'BASE': 650000},
]

# Suscripciones Familia
SUSCRIPCIONES_FAMILIA = [
    {'CONCEPTO': 'Giganet', 'CATEGORIA': 'Suscripciones', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 5, 'BASE': 250000},
    {'CONCEPTO': 'Tigo Familiar', 'CATEGORIA': 'Suscripciones', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 10, 'BASE': 320000},
]

# Variables Familia (presupuesto mensual)
VARIABLES_FAMILIA = [
    {'CONCEPTO': 'Supermercado', 'TIPO': 'Variable', 'MONTO': 2500000},
    {'CONCEPTO': 'Combustible', 'TIPO': 'Variable', 'MONTO': 800000},
    {'CONCEPTO': 'Salud', 'TIPO': 'Variable', 'MONTO': 500000},
    {'CONCEPTO': 'Recreacion', 'TIPO': 'Variable', 'MONTO': 400000},
    {'CONCEPTO': 'Varios', 'TIPO': 'Variable', 'MONTO': 300000},
    {'CONCEPTO': 'Reserva Variable', 'TIPO': 'Variable', 'MONTO': 0},
]

# Ingresos NeuroTEA
INGRESOS_NT = [
    {'CONCEPTO': 'Ingresos Terapias', 'TIPO': 'Variable/Mensual', 'MONTO': 25000000},
    {'CONCEPTO': 'Evaluaciones', 'TIPO': 'Variable', 'MONTO': 3000000},
    {'CONCEPTO': 'Talleres', 'TIPO': 'Variable', 'MONTO': 2000000},
    {'CONCEPTO': 'Reserva', 'TIPO': '', 'MONTO': 0},
]

# Gastos Clinica NT
GASTOS_CLINICA_NT = [
    {'CONCEPTO': 'Alquiler 1 (Principal)', 'CATEGORIA': 'Clinica', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 5, 'BASE': 13500000},
    {'CONCEPTO': 'Alquiler 2 (Secundario)', 'CATEGORIA': 'Clinica', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 5, 'BASE': 4130000},
    {'CONCEPTO': 'ANDE Clinica', 'CATEGORIA': 'Clinica', 'FRECUENCIA': 'Variable/Mensual', 'DIA': 15, 'BASE': 350000},
]

# Sueldos y Honorarios NT
SUELDOS_NT = [
    {'CONCEPTO': 'Sueldo Aracely', 'CATEGORIA': 'Sueldos', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 30, 'BASE': 2000000},
    {'CONCEPTO': 'Sueldo Fatima', 'CATEGORIA': 'Sueldos', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 30, 'BASE': 1800000},
    {'CONCEPTO': 'Honorario Contador', 'CATEGORIA': 'Sueldos', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 5, 'BASE': 500000},
    {'CONCEPTO': 'Salario Administrador', 'CATEGORIA': 'Sueldos', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 30, 'BASE': 5000000},
    {'CONCEPTO': 'Honorario Mant. Sistema', 'CATEGORIA': 'Sueldos', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 5, 'BASE': 300000},
]

# Telefonía NT
TELEFONIA_NT = [
    {'CONCEPTO': 'Celular Tigo NeuroTEA', 'CATEGORIA': 'Telefonia', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 10, 'BASE': 100000},
    {'CONCEPTO': 'Celular Tigo Sistema', 'CATEGORIA': 'Telefonia', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 10, 'BASE': 100000},
    {'CONCEPTO': 'WhatsFlow', 'CATEGORIA': 'Telefonia', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 5, 'BASE': 150000},
    {'CONCEPTO': 'Internet NeuroTEA', 'CATEGORIA': 'Telefonia', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 10, 'BASE': 200000},
]

# Obligaciones NT
OBLIGACIONES_NT = [
    {'CONCEPTO': 'IVA', 'CATEGORIA': 'Obligaciones', 'FRECUENCIA': 'Variable/Mensual', 'DIA': 20, 'BASE': 1500000},
    {'CONCEPTO': 'IPS', 'CATEGORIA': 'Obligaciones', 'FRECUENCIA': 'Fijo/Mensual', 'DIA': 15, 'BASE': 800000},
]

# Variables NT
VARIABLES_NT = [
    {'CONCEPTO': 'Materiales Terapia', 'TIPO': 'Variable', 'MONTO': 300000},
    {'CONCEPTO': 'Mantenimiento', 'TIPO': 'Variable', 'MONTO': 200000},
    {'CONCEPTO': 'Varios NT', 'TIPO': 'Variable', 'MONTO': 150000},
]

# Cuentas para saldos
CUENTAS_FAMILIA = ['ITAU Marco', 'Coop. Univ. Marco', 'ITAU Clara', 'UENO Clara', 'Tarjeta Solar', 'Tarjeta ITAU Clara', 'Gourmed']
CUENTAS_NT = ['Atlas NeuroTEA', 'Costos Operativos', 'Utilidad del Dueño', 'Fondo Emergencia', 'Fondo Inversión']


# ============================================================================
# ESTILOS
# ============================================================================
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def estilo_header(ws, cell, color):
    """Aplicar estilo de header"""
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

def estilo_seccion(ws, cell, color):
    """Aplicar estilo de sección"""
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')

def estilo_subseccion(ws, cell, color_claro):
    """Aplicar estilo de subsección"""
    cell.font = Font(bold=True, size=9)
    cell.fill = PatternFill(start_color=color_claro, end_color=color_claro, fill_type='solid')

def formato_numero(cell):
    """Formato de número con separador de miles"""
    cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal='right')


# ============================================================================
# CREAR HOJAS
# ============================================================================

def crear_hoja_configuracion(wb):
    """Crear hoja CONFIGURACION"""
    ws = wb.create_sheet("CONFIGURACION", 0)

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = 'CONFIGURACION DEL SISTEMA'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers de listas
    headers = ['ENTIDAD', 'TIPO MOV.', 'FRECUENCIA', 'ESTADO', 'CUENTAS', 'MESES', 'META', 'VALOR']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')

    # Datos
    row = 4
    max_rows = max(len(CONFIG['ENTIDADES']), len(CONFIG['TIPO_MOV']), len(CONFIG['FRECUENCIA']),
                   len(CONFIG['ESTADO']), len(CONFIG['CUENTAS']), len(CONFIG['MESES']))

    for i in range(max_rows):
        if i < len(CONFIG['ENTIDADES']):
            ws.cell(row=row+i, column=1, value=CONFIG['ENTIDADES'][i])
        if i < len(CONFIG['TIPO_MOV']):
            ws.cell(row=row+i, column=2, value=CONFIG['TIPO_MOV'][i])
        if i < len(CONFIG['FRECUENCIA']):
            ws.cell(row=row+i, column=3, value=CONFIG['FRECUENCIA'][i])
        if i < len(CONFIG['ESTADO']):
            ws.cell(row=row+i, column=4, value=CONFIG['ESTADO'][i])
        if i < len(CONFIG['CUENTAS']):
            ws.cell(row=row+i, column=5, value=CONFIG['CUENTAS'][i])
        if i < len(CONFIG['MESES']):
            ws.cell(row=row+i, column=6, value=CONFIG['MESES'][i])

    # Metas
    meta_row = 4
    for nombre, valor in CONFIG['META'].items():
        ws.cell(row=meta_row, column=7, value=nombre)
        ws.cell(row=meta_row, column=8, value=valor)
        ws.cell(row=meta_row, column=8).number_format = '0%'
        meta_row += 1

    # Categorías
    ws.cell(row=16, column=1, value='CATEGORIAS FAMILIA')
    ws.cell(row=16, column=2, value='CATEGORIAS NEUROTEA')
    ws.cell(row=16, column=1).font = Font(bold=True)
    ws.cell(row=16, column=2).font = Font(bold=True)

    for i, cat in enumerate(CONFIG['CATEGORIAS_FAMILIA']):
        ws.cell(row=17+i, column=1, value=cat)
    for i, cat in enumerate(CONFIG['CATEGORIAS_NT']):
        ws.cell(row=17+i, column=2, value=cat)

    # Ajustar anchos
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    return ws


def crear_hoja_tablero(wb):
    """Crear hoja TABLERO - Dashboard principal"""
    ws = wb.create_sheet("TABLERO", 1)

    # Configurar anchos
    anchos = {'A': 22, 'B': 14, 'C': 14, 'D': 14, 'E': 3, 'F': 22, 'G': 14, 'H': 14, 'I': 14}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    # ===== TÍTULO =====
    ws.merge_cells('A1:I1')
    ws['A1'] = 'TABLERO DE CONTROL FINANCIERO'
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color=COLORES['GRIS_HEADER'], end_color=COLORES['GRIS_HEADER'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Mes y Fecha
    ws['B2'] = 'MES:'
    ws['C2'] = 'Enero 2026'
    ws['C2'].font = Font(bold=True)
    ws['G2'] = 'Hoy:'
    ws['H2'] = datetime.now().strftime('%d/%m/%Y')

    row = 4

    # ===== HEADER FAMILIA =====
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'FAMILIA'
    estilo_header(ws, ws[f'A{row}'], COLORES['VERDE_HEADER'])
    ws.row_dimensions[row].height = 25

    # ===== HEADER NEUROTEA =====
    ws.merge_cells(f'F{row}:I{row}')
    ws[f'F{row}'] = 'NEUROTEA'
    estilo_header(ws, ws[f'F{row}'], COLORES['AZUL_HEADER'])

    row += 2

    # ===== SALDOS EN CUENTAS - FAMILIA =====
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'SALDOS EN CUENTAS'
    estilo_subseccion(ws, ws[f'A{row}'], COLORES['VERDE_CLARO'])

    row += 1
    # Headers tabla saldos
    for col, header in [('A', 'Cuenta'), ('B', 'Esperado'), ('C', 'Real'), ('D', 'Diferencia')]:
        cell = ws[f'{col}{row}']
        cell.value = header
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = thin_border

    row += 1
    start_saldos_fam = row
    for cuenta in CUENTAS_FAMILIA:
        ws[f'A{row}'] = cuenta
        ws[f'B{row}'] = 0
        ws[f'C{row}'] = 0
        ws[f'D{row}'] = f'=C{row}-B{row}'
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = thin_border
            if col in ['B', 'C', 'D']:
                formato_numero(ws[f'{col}{row}'])
        row += 1

    # Total saldos familia
    ws[f'A{row}'] = 'TOTAL'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f'=SUM(B{start_saldos_fam}:B{row-1})'
    ws[f'C{row}'] = f'=SUM(C{start_saldos_fam}:C{row-1})'
    ws[f'D{row}'] = f'=C{row}-B{row}'
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')
        if col in ['B', 'C', 'D']:
            formato_numero(ws[f'{col}{row}'])

    # ===== NEUROTEA - INDICADORES DE METAS =====
    row_nt = 6
    ws.merge_cells(f'F{row_nt}:I{row_nt}')
    ws[f'F{row_nt}'] = 'INDICADORES DE METAS'
    estilo_subseccion(ws, ws[f'F{row_nt}'], COLORES['AZUL_CLARO'])

    row_nt += 1
    # Ingresos del mes
    ws[f'F{row_nt}'] = 'INGRESOS DEL MES'
    ws[f'F{row_nt}'].font = Font(size=9, color='059669')
    ws[f'H{row_nt}'] = 'GASTOS DEL MES'
    ws[f'H{row_nt}'].font = Font(size=9, color='DC2626')

    row_nt += 1
    ws[f'F{row_nt}'] = 30000000
    ws[f'F{row_nt}'].font = Font(bold=True, size=14, color='059669')
    formato_numero(ws[f'F{row_nt}'])
    ws[f'H{row_nt}'] = 27300000
    ws[f'H{row_nt}'].font = Font(bold=True, size=14, color='1F2937')
    formato_numero(ws[f'H{row_nt}'])

    row_nt += 1
    ws[f'F{row_nt}'] = 'GANANCIA REAL'
    ws[f'F{row_nt}'].font = Font(size=9, color='059669')
    ws[f'H{row_nt}'] = 'META 7%'
    ws[f'H{row_nt}'].font = Font(size=9, color='1E40AF')

    row_nt += 1
    ws[f'F{row_nt}'] = 2700000
    ws[f'F{row_nt}'].font = Font(bold=True, size=14, color='059669')
    formato_numero(ws[f'F{row_nt}'])
    ws[f'H{row_nt}'] = 2100000
    ws[f'H{row_nt}'].font = Font(bold=True, size=14, color='1E40AF')
    formato_numero(ws[f'H{row_nt}'])

    row_nt += 2
    ws[f'F{row_nt}'] = '% Gastos sobre Ingresos'
    ws[f'I{row_nt}'] = '91% / 93% máx'
    ws[f'I{row_nt}'].font = Font(bold=True)

    row_nt += 2
    # Distribución ganancia
    ws[f'F{row_nt}'] = 'Distribución de Ganancia (7%)'
    ws[f'F{row_nt}'].font = Font(bold=True, size=9)

    row_nt += 1
    for col, (titulo, monto) in [('F', ('Utilidad Dueño', 700000)),
                                   ('G', ('Fondo Emerg.', 700000)),
                                   ('H', ('Fondo Inversión', 700000))]:
        ws[f'{col}{row_nt}'] = titulo
        ws[f'{col}{row_nt}'].font = Font(size=8)
        ws[f'{col}{row_nt+1}'] = monto
        ws[f'{col}{row_nt+1}'].font = Font(bold=True)
        formato_numero(ws[f'{col}{row_nt+1}'])

    # ===== SALDOS NEUROTEA =====
    row_nt += 4
    ws.merge_cells(f'F{row_nt}:I{row_nt}')
    ws[f'F{row_nt}'] = 'SALDOS EN CUENTAS'
    estilo_subseccion(ws, ws[f'F{row_nt}'], COLORES['AZUL_CLARO'])

    row_nt += 1
    for col, header in [('F', 'Cuenta'), ('G', 'Mes'), ('H', 'Acumulado')]:
        cell = ws[f'{col}{row_nt}']
        cell.value = header
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = thin_border

    row_nt += 1
    for cuenta in CUENTAS_NT:
        ws[f'F{row_nt}'] = cuenta
        ws[f'G{row_nt}'] = 0
        ws[f'H{row_nt}'] = 0
        for col in ['F', 'G', 'H']:
            ws[f'{col}{row_nt}'].border = thin_border
            if col in ['G', 'H']:
                formato_numero(ws[f'{col}{row_nt}'])
        row_nt += 1

    row = max(row, row_nt) + 2

    # ===== PRESUPUESTO VS REAL - FAMILIA =====
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'PRESUPUESTO vs REAL'
    estilo_subseccion(ws, ws[f'A{row}'], COLORES['VERDE_CLARO'])

    # Headers
    row += 1
    for col, header in [('A', 'Categoría'), ('B', 'Presupuesto'), ('C', 'Real'), ('D', 'Estado')]:
        cell = ws[f'{col}{row}']
        cell.value = header
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = thin_border

    row += 1
    categorias_fam = [
        ('INGRESOS FAMILIA', 15200000, 14500000),
        ('GASTOS FIJOS', 7250000, 7100000),
        ('CUOTAS Y PRÉSTAMOS', 5541000, 5541000),
        ('OBLIGACIONES LEGALES', 450000, 380000),
        ('SUSCRIPCIONES', 520000, 520000),
        ('VARIABLES', 900000, 1250000),
        ('AHORRO', 500000, 0),
    ]

    for cat, ppto, real in categorias_fam:
        ws[f'A{row}'] = f'► {cat}'
        ws[f'B{row}'] = ppto
        ws[f'C{row}'] = real
        pct = real/ppto if ppto > 0 else 0
        ws[f'D{row}'] = f'{int(pct*100)}%'
        ws[f'D{row}'].alignment = Alignment(horizontal='center')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = thin_border
            if col in ['B', 'C']:
                formato_numero(ws[f'{col}{row}'])
        row += 1

    # Balance Familia
    ws[f'A{row}'] = 'BALANCE FAMILIA'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = 39000
    ws[f'C{row}'] = -291000
    ws[f'C{row}'].font = Font(color='DC2626')
    ws[f'D{row}'] = 'DÉFICIT'
    ws[f'D{row}'].font = Font(bold=True, color='FFFFFF', size=8)
    ws[f'D{row}'].fill = PatternFill(start_color=COLORES['ROJO'], end_color=COLORES['ROJO'], fill_type='solid')
    ws[f'D{row}'].alignment = Alignment(horizontal='center')
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].border = thin_border
        if col in ['B', 'C']:
            formato_numero(ws[f'{col}{row}'])

    row += 2

    # ===== FLUJO DEL MES - FAMILIA =====
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'FLUJO DEL MES'
    estilo_subseccion(ws, ws[f'A{row}'], COLORES['VERDE_CLARO'])

    row += 1
    flujo_fam = [
        ('Ingresos', 14500000, '059669'),
        ('Egresos Pagados', -12450000, 'DC2626'),
        ('Egresos Pendientes', -2341000, 'F59E0B'),
    ]

    for concepto, monto, color in flujo_fam:
        ws[f'A{row}'] = concepto
        ws[f'D{row}'] = monto
        ws[f'D{row}'].font = Font(color=color)
        formato_numero(ws[f'D{row}'])
        row += 1

    # Balance flujo
    ws[f'A{row}'] = 'BALANCE'
    ws[f'A{row}'].font = Font(bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['GRIS_HEADER'], end_color=COLORES['GRIS_HEADER'], fill_type='solid')
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'D{row}'] = 2050000
    ws[f'D{row}'].font = Font(bold=True, color='FFFFFF')
    ws[f'D{row}'].fill = PatternFill(start_color=COLORES['GRIS_HEADER'], end_color=COLORES['GRIS_HEADER'], fill_type='solid')
    formato_numero(ws[f'D{row}'])

    row += 2

    # ===== LIQUIDEZ PRÓXIMOS PAGOS - FAMILIA =====
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'LIQUIDEZ - PRÓXIMOS PAGOS'
    estilo_subseccion(ws, ws[f'A{row}'], COLORES['VERDE_CLARO'])

    row += 1
    for col, header in [('A', 'Concepto'), ('B', 'Cuotas'), ('C', 'Monto'), ('D', 'Saldo')]:
        cell = ws[f'{col}{row}']
        cell.value = header
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = thin_border

    row += 1
    liquidez_data = [
        ('Caja disponible', '-', '-', 2350000, ''),
        ('Atrasados', 2, -850000, 1500000, 'PAGAR'),
        ('Esta semana', 4, -2100000, -600000, 'FALTA'),
        ('Próxima semana', 3, -1800000, -2400000, 'FALTA'),
        ('3ra semana', 2, -1200000, -3600000, 'FALTA'),
    ]

    for concepto, cuotas, monto, saldo, estado in liquidez_data:
        ws[f'A{row}'] = concepto
        ws[f'B{row}'] = cuotas if cuotas != '-' else '-'
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        ws[f'C{row}'] = monto if monto != '-' else '-'
        ws[f'D{row}'] = saldo
        if estado:
            # Agregar columna E para estado
            ws[f'E{row}'] = estado
            if estado == 'PAGAR':
                ws[f'E{row}'].fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            elif estado == 'FALTA':
                ws[f'E{row}'].fill = PatternFill(start_color=COLORES['ROJO_CLARO'], end_color=COLORES['ROJO_CLARO'], fill_type='solid')
            ws[f'E{row}'].font = Font(bold=True, size=8)
            ws[f'E{row}'].alignment = Alignment(horizontal='center')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = thin_border
            if col in ['C', 'D']:
                formato_numero(ws[f'{col}{row}'])
        row += 1

    # Saldo final
    ws[f'A{row}'] = 'SALDO FINAL'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = 11
    ws[f'B{row}'].alignment = Alignment(horizontal='center')
    ws[f'C{row}'] = -5950000
    ws[f'D{row}'] = -3600000
    ws[f'E{row}'] = 'DÉFICIT'
    ws[f'E{row}'].fill = PatternFill(start_color=COLORES['ROJO'], end_color=COLORES['ROJO'], fill_type='solid')
    ws[f'E{row}'].font = Font(bold=True, size=8, color='FFFFFF')
    ws[f'E{row}'].alignment = Alignment(horizontal='center')
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        if col in ['C', 'D']:
            formato_numero(ws[f'{col}{row}'])

    row_final_fam = row

    # ===== PRESUPUESTO VS REAL - NEUROTEA =====
    row_nt = 28
    ws.merge_cells(f'F{row_nt}:I{row_nt}')
    ws[f'F{row_nt}'] = 'PRESUPUESTO vs REAL'
    estilo_subseccion(ws, ws[f'F{row_nt}'], COLORES['AZUL_CLARO'])

    row_nt += 1
    for col, header in [('F', 'Categoría'), ('G', 'Presupuesto'), ('H', 'Real'), ('I', 'Estado')]:
        cell = ws[f'{col}{row_nt}']
        cell.value = header
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = thin_border

    row_nt += 1
    categorias_nt = [
        ('INGRESOS NT', 30000000, 30000000),
        ('CLÍNICA', 17630000, 17630000),
        ('SUELDOS Y HONORARIOS', 9600000, 9300000),
        ('TELEFONÍA E INTERNET', 550000, 550000),
        ('OBLIGACIONES LEGALES', 2300000, 1850000),
        ('EVENTOS', 0, 0),
        ('VARIABLES', 450000, 670000),
        ('GANANCIA (7%)', 2100000, 2700000),
    ]

    for cat, ppto, real in categorias_nt:
        ws[f'F{row_nt}'] = f'► {cat}'
        ws[f'G{row_nt}'] = ppto
        ws[f'H{row_nt}'] = real
        pct = real/ppto if ppto > 0 else 0
        ws[f'I{row_nt}'] = f'{int(pct*100)}%' if ppto > 0 else '-%'
        ws[f'I{row_nt}'].alignment = Alignment(horizontal='center')
        for col in ['F', 'G', 'H', 'I']:
            ws[f'{col}{row_nt}'].border = thin_border
            if col in ['G', 'H']:
                formato_numero(ws[f'{col}{row_nt}'])
        row_nt += 1

    # Balance NT
    ws[f'F{row_nt}'] = 'BALANCE NEUROTEA'
    ws[f'F{row_nt}'].font = Font(bold=True)
    ws[f'G{row_nt}'] = -530000
    ws[f'H{row_nt}'] = 2700000
    ws[f'H{row_nt}'].font = Font(color='059669')
    ws[f'I{row_nt}'] = 'SUPERÁVIT'
    ws[f'I{row_nt}'].font = Font(bold=True, color='FFFFFF', size=8)
    ws[f'I{row_nt}'].fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    ws[f'I{row_nt}'].alignment = Alignment(horizontal='center')
    for col in ['F', 'G', 'H', 'I']:
        ws[f'{col}{row_nt}'].border = thin_border
        if col in ['G', 'H']:
            formato_numero(ws[f'{col}{row_nt}'])

    row_nt += 2

    # ===== FLUJO DEL MES - NEUROTEA =====
    ws.merge_cells(f'F{row_nt}:I{row_nt}')
    ws[f'F{row_nt}'] = 'FLUJO DEL MES'
    estilo_subseccion(ws, ws[f'F{row_nt}'], COLORES['AZUL_CLARO'])

    row_nt += 1
    flujo_nt = [
        ('Ingresos', 30000000, '059669'),
        ('Egresos Pagados', -24500000, 'DC2626'),
        ('Egresos Pendientes', -2800000, 'F59E0B'),
    ]

    for concepto, monto, color in flujo_nt:
        ws[f'F{row_nt}'] = concepto
        ws[f'I{row_nt}'] = monto
        ws[f'I{row_nt}'].font = Font(color=color)
        formato_numero(ws[f'I{row_nt}'])
        row_nt += 1

    ws[f'F{row_nt}'] = 'BALANCE'
    ws[f'F{row_nt}'].font = Font(bold=True, color='FFFFFF')
    ws[f'F{row_nt}'].fill = PatternFill(start_color=COLORES['GRIS_HEADER'], end_color=COLORES['GRIS_HEADER'], fill_type='solid')
    ws.merge_cells(f'F{row_nt}:H{row_nt}')
    ws[f'I{row_nt}'] = 5500000
    ws[f'I{row_nt}'].font = Font(bold=True, color='FFFFFF')
    ws[f'I{row_nt}'].fill = PatternFill(start_color=COLORES['GRIS_HEADER'], end_color=COLORES['GRIS_HEADER'], fill_type='solid')
    formato_numero(ws[f'I{row_nt}'])

    # ===== BALANCE CRUZADO =====
    row = max(row_final_fam, row_nt) + 3

    ws.merge_cells(f'A{row}:I{row}')
    ws[f'A{row}'] = 'BALANCE CRUZADO: NEUROTEA ↔ FAMILIA'
    ws[f'A{row}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['MORADO'], end_color=COLORES['MORADO'], fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 25

    row += 1
    for col, header in [('A', 'Concepto'), ('C', 'Este Mes'), ('E', 'Acumulado Año')]:
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = 'Préstamo NT → Familia'
    ws[f'C{row}'] = 3000000
    ws[f'E{row}'] = 8500000
    formato_numero(ws[f'C{row}'])
    formato_numero(ws[f'E{row}'])
    ws[f'C{row}'].font = Font(color='1E40AF')
    ws[f'E{row}'].font = Font(color='1E40AF')

    row += 1
    ws[f'A{row}'] = 'Devolución Familia → NT'
    ws[f'C{row}'] = '-'
    ws[f'E{row}'] = 2000000
    formato_numero(ws[f'E{row}'])

    row += 1
    ws[f'A{row}'] = 'SALDO NETO'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = 3000000
    ws[f'E{row}'] = 6500000
    formato_numero(ws[f'C{row}'])
    formato_numero(ws[f'E{row}'])
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['MORADO_CLARO'], end_color=COLORES['MORADO_CLARO'], fill_type='solid')
    ws[f'C{row}'].fill = PatternFill(start_color=COLORES['MORADO_CLARO'], end_color=COLORES['MORADO_CLARO'], fill_type='solid')
    ws[f'E{row}'].fill = PatternFill(start_color=COLORES['MORADO_CLARO'], end_color=COLORES['MORADO_CLARO'], fill_type='solid')

    # Alerta
    row += 2
    ws.merge_cells(f'F{row}:I{row+2}')
    ws[f'F{row}'] = 'NT SUBSIDIA A FAMILIA\nGs. 6.500.000'
    ws[f'F{row}'].font = Font(bold=True, size=11, color='DC2626')
    ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[f'F{row}'].fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

    return ws


def crear_hoja_presupuesto(wb):
    """Crear hoja PRESUPUESTO anual"""
    ws = wb.create_sheet("PRESUPUESTO", 2)

    # Anchos
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 15
    for i, mes in enumerate(CONFIG['MESES_CORTO']):
        ws.column_dimensions[get_column_letter(3+i)].width = 12
    ws.column_dimensions['O'].width = 14  # Total anual

    row = 1

    # Título
    ws.merge_cells('A1:O1')
    ws['A1'] = 'PRESUPUESTO ANUAL 2026'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = 'Los valores en azul son editables'
    ws['A2'].font = Font(italic=True, color='1E40AF')

    row = 4

    # ===== PRESUPUESTO FAMILIA =====
    ws.merge_cells(f'A{row}:O{row}')
    ws[f'A{row}'] = 'PRESUPUESTO FAMILIA'
    estilo_header(ws, ws[f'A{row}'], COLORES['VERDE_HEADER'])

    row += 1
    # Headers
    headers = ['CONCEPTO', 'TIPO'] + CONFIG['MESES_CORTO'] + ['TOTAL']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = thin_border

    row += 1

    # INGRESOS FAMILIA
    ws[f'A{row}'] = 'INGRESOS FAMILIA'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')
    row += 1

    start_ing = row
    for item in INGRESOS_FAMILIA:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value=item['TIPO']).border = thin_border
        for col in range(3, 15):
            ws.cell(row=row, column=col, value=item['MONTO']).border = thin_border
            formato_numero(ws.cell(row=row, column=col))
        # Total anual
        ws.cell(row=row, column=15, value=f'=SUM(C{row}:N{row})').border = thin_border
        formato_numero(ws.cell(row=row, column=15))
        row += 1

    # Total ingresos
    ws[f'A{row}'] = 'TOTAL INGRESOS FAMILIA'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_ing}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')
        formato_numero(ws.cell(row=row, column=col))
    total_ing_row = row

    row += 1

    # EGRESOS - GASTOS FIJOS
    ws[f'A{row}'] = 'EGRESOS - GASTOS FIJOS'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['ROJO_CLARO'], end_color=COLORES['ROJO_CLARO'], fill_type='solid')
    row += 1

    start_gf = row
    for item in GASTOS_FIJOS_FAMILIA:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value=item['FRECUENCIA']).border = thin_border
        for col in range(3, 15):
            ws.cell(row=row, column=col, value=item['BASE']).border = thin_border
            formato_numero(ws.cell(row=row, column=col))
        ws.cell(row=row, column=15, value=f'=SUM(C{row}:N{row})').border = thin_border
        formato_numero(ws.cell(row=row, column=15))
        row += 1

    ws[f'A{row}'] = 'SUBTOTAL GASTOS FIJOS'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_gf}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = thin_border
        formato_numero(ws.cell(row=row, column=col))
    subtotal_gf_row = row

    row += 1

    # EGRESOS - CUOTAS Y PRESTAMOS
    ws[f'A{row}'] = 'EGRESOS - CUOTAS Y PRESTAMOS'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['ROJO_CLARO'], end_color=COLORES['ROJO_CLARO'], fill_type='solid')
    row += 1

    start_cuotas = row
    for item in CUOTAS_FAMILIA:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value=item['FRECUENCIA']).border = thin_border
        for col in range(3, 15):
            ws.cell(row=row, column=col, value=item['BASE']).border = thin_border
            formato_numero(ws.cell(row=row, column=col))
        ws.cell(row=row, column=15, value=f'=SUM(C{row}:N{row})').border = thin_border
        formato_numero(ws.cell(row=row, column=15))
        row += 1

    ws[f'A{row}'] = 'SUBTOTAL CUOTAS Y PRESTAMOS'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_cuotas}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = thin_border
        formato_numero(ws.cell(row=row, column=col))
    subtotal_cuotas_row = row

    row += 1

    # EGRESOS - SUSCRIPCIONES
    ws[f'A{row}'] = 'EGRESOS - SUSCRIPCIONES'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    start_susc = row
    for item in SUSCRIPCIONES_FAMILIA:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value=item['FRECUENCIA']).border = thin_border
        for col in range(3, 15):
            ws.cell(row=row, column=col, value=item['BASE']).border = thin_border
            formato_numero(ws.cell(row=row, column=col))
        ws.cell(row=row, column=15, value=f'=SUM(C{row}:N{row})').border = thin_border
        formato_numero(ws.cell(row=row, column=15))
        row += 1

    ws[f'A{row}'] = 'SUBTOTAL SUSCRIPCIONES'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_susc}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = thin_border
        formato_numero(ws.cell(row=row, column=col))
    subtotal_susc_row = row

    row += 1

    # EGRESOS - VARIABLES
    ws[f'A{row}'] = 'EGRESOS - VARIABLES'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    start_var = row
    for item in VARIABLES_FAMILIA:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value=item['TIPO']).border = thin_border
        for col in range(3, 15):
            ws.cell(row=row, column=col, value=item['MONTO']).border = thin_border
            formato_numero(ws.cell(row=row, column=col))
        ws.cell(row=row, column=15, value=f'=SUM(C{row}:N{row})').border = thin_border
        formato_numero(ws.cell(row=row, column=15))
        row += 1

    ws[f'A{row}'] = 'SUBTOTAL VARIABLES'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_var}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = thin_border
        formato_numero(ws.cell(row=row, column=col))
    subtotal_var_row = row

    row += 1

    # TOTAL EGRESOS
    ws[f'A{row}'] = 'TOTAL EGRESOS FAMILIA'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={col_letter}{subtotal_gf_row}+{col_letter}{subtotal_cuotas_row}+{col_letter}{subtotal_susc_row}+{col_letter}{subtotal_var_row}')
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORES['ROJO_CLARO'], end_color=COLORES['ROJO_CLARO'], fill_type='solid')
        formato_numero(ws.cell(row=row, column=col))
    total_egr_row = row

    row += 1

    # BALANCE FAMILIA
    ws[f'A{row}'] = 'BALANCE FAMILIA (Ing - Egr)'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={col_letter}{total_ing_row}-{col_letter}{total_egr_row}')
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        ws.cell(row=row, column=col).font = Font(bold=True)
        formato_numero(ws.cell(row=row, column=col))

    row += 3

    # ===== PRESUPUESTO NEUROTEA =====
    ws.merge_cells(f'A{row}:O{row}')
    ws[f'A{row}'] = 'PRESUPUESTO NEUROTEA'
    estilo_header(ws, ws[f'A{row}'], COLORES['AZUL_HEADER'])

    row += 1
    # Headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = thin_border

    row += 1

    # INGRESOS NT
    ws[f'A{row}'] = 'INGRESOS NEUROTEA'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['AZUL_CLARO'], end_color=COLORES['AZUL_CLARO'], fill_type='solid')
    row += 1

    start_ing_nt = row
    for item in INGRESOS_NT:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value=item['TIPO']).border = thin_border
        for col in range(3, 15):
            ws.cell(row=row, column=col, value=item['MONTO']).border = thin_border
            formato_numero(ws.cell(row=row, column=col))
        ws.cell(row=row, column=15, value=f'=SUM(C{row}:N{row})').border = thin_border
        formato_numero(ws.cell(row=row, column=15))
        row += 1

    ws[f'A{row}'] = 'TOTAL INGRESOS NEUROTEA'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_ing_nt}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORES['AZUL_CLARO'], end_color=COLORES['AZUL_CLARO'], fill_type='solid')
        formato_numero(ws.cell(row=row, column=col))
    total_ing_nt_row = row

    row += 1

    # EGRESOS NT - CLINICA
    ws[f'A{row}'] = 'EGRESOS NT - CLINICA'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    start_clinica = row
    for item in GASTOS_CLINICA_NT:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value=item['FRECUENCIA']).border = thin_border
        for col in range(3, 15):
            ws.cell(row=row, column=col, value=item['BASE']).border = thin_border
            formato_numero(ws.cell(row=row, column=col))
        ws.cell(row=row, column=15, value=f'=SUM(C{row}:N{row})').border = thin_border
        formato_numero(ws.cell(row=row, column=15))
        row += 1

    ws[f'A{row}'] = 'SUBTOTAL CLINICA'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_clinica}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = thin_border
        formato_numero(ws.cell(row=row, column=col))
    subtotal_clinica_row = row

    row += 1

    # EGRESOS NT - SUELDOS
    ws[f'A{row}'] = 'EGRESOS NT - SUELDOS Y HONORARIOS'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    start_sueldos = row
    for item in SUELDOS_NT:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value=item['FRECUENCIA']).border = thin_border
        for col in range(3, 15):
            ws.cell(row=row, column=col, value=item['BASE']).border = thin_border
            formato_numero(ws.cell(row=row, column=col))
        ws.cell(row=row, column=15, value=f'=SUM(C{row}:N{row})').border = thin_border
        formato_numero(ws.cell(row=row, column=15))
        row += 1

    ws[f'A{row}'] = 'SUBTOTAL SUELDOS'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_sueldos}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = thin_border
        formato_numero(ws.cell(row=row, column=col))
    subtotal_sueldos_row = row

    row += 1

    # EGRESOS NT - TELEFONIA
    ws[f'A{row}'] = 'EGRESOS NT - TELEFONIA E INTERNET'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    start_tel = row
    for item in TELEFONIA_NT:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value=item['FRECUENCIA']).border = thin_border
        for col in range(3, 15):
            ws.cell(row=row, column=col, value=item['BASE']).border = thin_border
            formato_numero(ws.cell(row=row, column=col))
        ws.cell(row=row, column=15, value=f'=SUM(C{row}:N{row})').border = thin_border
        formato_numero(ws.cell(row=row, column=15))
        row += 1

    ws[f'A{row}'] = 'SUBTOTAL TELEFONIA'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_tel}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = thin_border
        formato_numero(ws.cell(row=row, column=col))
    subtotal_tel_row = row

    row += 1

    # EGRESOS NT - OBLIGACIONES
    ws[f'A{row}'] = 'EGRESOS NT - OBLIGACIONES LEGALES'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    start_obl = row
    for item in OBLIGACIONES_NT:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value=item['FRECUENCIA']).border = thin_border
        for col in range(3, 15):
            ws.cell(row=row, column=col, value=item['BASE']).border = thin_border
            formato_numero(ws.cell(row=row, column=col))
        ws.cell(row=row, column=15, value=f'=SUM(C{row}:N{row})').border = thin_border
        formato_numero(ws.cell(row=row, column=15))
        row += 1

    ws[f'A{row}'] = 'SUBTOTAL OBLIGACIONES'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_obl}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = thin_border
        formato_numero(ws.cell(row=row, column=col))
    subtotal_obl_row = row

    row += 1

    # EGRESOS NT - VARIABLES
    ws[f'A{row}'] = 'EGRESOS NT - VARIABLES'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    start_var_nt = row
    for item in VARIABLES_NT:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value=item['TIPO']).border = thin_border
        for col in range(3, 15):
            ws.cell(row=row, column=col, value=item['MONTO']).border = thin_border
            formato_numero(ws.cell(row=row, column=col))
        ws.cell(row=row, column=15, value=f'=SUM(C{row}:N{row})').border = thin_border
        formato_numero(ws.cell(row=row, column=15))
        row += 1

    ws[f'A{row}'] = 'SUBTOTAL VARIABLES NT'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{start_var_nt}:{col_letter}{row-1})')
        ws.cell(row=row, column=col).border = thin_border
        formato_numero(ws.cell(row=row, column=col))
    subtotal_var_nt_row = row

    row += 1

    # TOTAL EGRESOS NT
    ws[f'A{row}'] = 'TOTAL EGRESOS NEUROTEA'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        formula = f'={col_letter}{subtotal_clinica_row}+{col_letter}{subtotal_sueldos_row}+{col_letter}{subtotal_tel_row}+{col_letter}{subtotal_obl_row}+{col_letter}{subtotal_var_nt_row}'
        ws.cell(row=row, column=col, value=formula)
        ws.cell(row=row, column=col).border = thin_border
        formato_numero(ws.cell(row=row, column=col))
    total_egr_nt_row = row

    row += 1

    # GANANCIA NT (META 7%)
    ws[f'A{row}'] = 'GANANCIA NEUROTEA (Meta 7%)'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={col_letter}{total_ing_nt_row}*0.07')
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        formato_numero(ws.cell(row=row, column=col))
    ganancia_row = row

    row += 1

    # BALANCE NT
    ws[f'A{row}'] = 'BALANCE NEUROTEA'
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(3, 16):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={col_letter}{total_ing_nt_row}-{col_letter}{total_egr_nt_row}-{col_letter}{ganancia_row}')
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORES['AZUL_CLARO'], end_color=COLORES['AZUL_CLARO'], fill_type='solid')
        ws.cell(row=row, column=col).font = Font(bold=True)
        formato_numero(ws.cell(row=row, column=col))

    return ws


def crear_hoja_gastos_fijos(wb):
    """Crear hoja GASTOS FIJOS"""
    ws = wb.create_sheet("GASTOS FIJOS", 3)

    # Anchos
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 14
    for i in range(12):
        ws.column_dimensions[get_column_letter(7+i)].width = 11

    row = 1

    # Título
    ws.merge_cells('A1:R1')
    ws['A1'] = 'GASTOS FIJOS - LISTA MAESTRA'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A2'] = 'DIA = día del mes que vence. BASE = monto base mensual (editable por mes si varía)'
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    # Headers
    headers = ['CONCEPTO', 'ENTIDAD', 'CATEGORIA', 'FRECUENCIA', 'DIA', 'BASE'] + CONFIG['MESES_CORTO']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = thin_border

    row += 1

    # GASTOS FIJOS FAMILIA
    ws[f'A{row}'] = 'GASTOS FIJOS FAMILIA'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')
    row += 1

    for item in GASTOS_FIJOS_FAMILIA:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value='Familia').border = thin_border
        ws.cell(row=row, column=3, value=item['CATEGORIA']).border = thin_border
        ws.cell(row=row, column=4, value=item['FRECUENCIA']).border = thin_border
        ws.cell(row=row, column=5, value=item['DIA']).border = thin_border
        ws.cell(row=row, column=6, value=item['BASE']).border = thin_border
        formato_numero(ws.cell(row=row, column=6))
        for col in range(7, 19):
            ws.cell(row=row, column=col, value='').border = thin_border
        row += 1

    # CUOTAS Y PRESTAMOS
    ws[f'A{row}'] = 'CUOTAS Y PRESTAMOS FAMILIA'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')
    row += 1

    for item in CUOTAS_FAMILIA:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value='Familia').border = thin_border
        ws.cell(row=row, column=3, value=item['CATEGORIA']).border = thin_border
        ws.cell(row=row, column=4, value=item['FRECUENCIA']).border = thin_border
        ws.cell(row=row, column=5, value=item['DIA']).border = thin_border
        ws.cell(row=row, column=6, value=item['BASE']).border = thin_border
        formato_numero(ws.cell(row=row, column=6))
        for col in range(7, 19):
            ws.cell(row=row, column=col, value='').border = thin_border
        row += 1

    # SUSCRIPCIONES
    ws[f'A{row}'] = 'SUSCRIPCIONES FAMILIA'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    for item in SUSCRIPCIONES_FAMILIA:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value='Familia').border = thin_border
        ws.cell(row=row, column=3, value=item['CATEGORIA']).border = thin_border
        ws.cell(row=row, column=4, value=item['FRECUENCIA']).border = thin_border
        ws.cell(row=row, column=5, value=item['DIA']).border = thin_border
        ws.cell(row=row, column=6, value=item['BASE']).border = thin_border
        formato_numero(ws.cell(row=row, column=6))
        for col in range(7, 19):
            ws.cell(row=row, column=col, value='').border = thin_border
        row += 1

    # CLINICA NT
    ws[f'A{row}'] = 'CLINICA NEUROTEA'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['AZUL_CLARO'], end_color=COLORES['AZUL_CLARO'], fill_type='solid')
    row += 1

    for item in GASTOS_CLINICA_NT:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value='NeuroTEA').border = thin_border
        ws.cell(row=row, column=3, value=item['CATEGORIA']).border = thin_border
        ws.cell(row=row, column=4, value=item['FRECUENCIA']).border = thin_border
        ws.cell(row=row, column=5, value=item['DIA']).border = thin_border
        ws.cell(row=row, column=6, value=item['BASE']).border = thin_border
        formato_numero(ws.cell(row=row, column=6))
        for col in range(7, 19):
            ws.cell(row=row, column=col, value='').border = thin_border
        row += 1

    # SUELDOS NT
    ws[f'A{row}'] = 'SUELDOS Y HONORARIOS NT'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['AZUL_CLARO'], end_color=COLORES['AZUL_CLARO'], fill_type='solid')
    row += 1

    for item in SUELDOS_NT:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value='NeuroTEA').border = thin_border
        ws.cell(row=row, column=3, value=item['CATEGORIA']).border = thin_border
        ws.cell(row=row, column=4, value=item['FRECUENCIA']).border = thin_border
        ws.cell(row=row, column=5, value=item['DIA']).border = thin_border
        ws.cell(row=row, column=6, value=item['BASE']).border = thin_border
        formato_numero(ws.cell(row=row, column=6))
        for col in range(7, 19):
            ws.cell(row=row, column=col, value='').border = thin_border
        row += 1

    # TELEFONIA NT
    ws[f'A{row}'] = 'TELEFONIA E INTERNET NT'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    for item in TELEFONIA_NT:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value='NeuroTEA').border = thin_border
        ws.cell(row=row, column=3, value=item['CATEGORIA']).border = thin_border
        ws.cell(row=row, column=4, value=item['FRECUENCIA']).border = thin_border
        ws.cell(row=row, column=5, value=item['DIA']).border = thin_border
        ws.cell(row=row, column=6, value=item['BASE']).border = thin_border
        formato_numero(ws.cell(row=row, column=6))
        for col in range(7, 19):
            ws.cell(row=row, column=col, value='').border = thin_border
        row += 1

    # OBLIGACIONES NT
    ws[f'A{row}'] = 'OBLIGACIONES LEGALES NT'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    for item in OBLIGACIONES_NT:
        ws.cell(row=row, column=1, value=item['CONCEPTO']).border = thin_border
        ws.cell(row=row, column=2, value='NeuroTEA').border = thin_border
        ws.cell(row=row, column=3, value=item['CATEGORIA']).border = thin_border
        ws.cell(row=row, column=4, value=item['FRECUENCIA']).border = thin_border
        ws.cell(row=row, column=5, value=item['DIA']).border = thin_border
        ws.cell(row=row, column=6, value=item['BASE']).border = thin_border
        formato_numero(ws.cell(row=row, column=6))
        for col in range(7, 19):
            ws.cell(row=row, column=col, value='').border = thin_border
        row += 1

    return ws


def crear_hoja_carga(wb):
    """Crear hoja CARGA para ingreso de datos"""
    ws = wb.create_sheet("CARGA", 4)

    # Anchos
    anchos = {'A': 12, 'B': 12, 'C': 18, 'D': 25, 'E': 30, 'F': 14, 'G': 15, 'H': 12}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = 'REGISTRO DE MOVIMIENTOS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = 'Solo cargar movimientos variables. Los fijos se cargan automáticamente.'
    ws['A2'].font = Font(italic=True, size=9)

    ws['A3'] = 'MES:'
    ws['B3'] = 'Enero'
    ws['B3'].font = Font(bold=True)

    # Validación para mes
    dv_mes = DataValidation(type="list", formula1='"' + ','.join(CONFIG['MESES']) + '"')
    ws.add_data_validation(dv_mes)
    dv_mes.add('B3')

    # Headers
    headers = ['FECHA', 'ENTIDAD', 'CATEGORIA', 'CONCEPTO', 'DESCRIPCION', 'MONTO', 'CUENTA', 'ESTADO']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORES['GRIS_HEADER'], end_color=COLORES['GRIS_HEADER'], fill_type='solid')
        cell.border = thin_border

    # Validaciones
    dv_entidad = DataValidation(type="list", formula1='"Familia,NeuroTEA"')
    ws.add_data_validation(dv_entidad)
    dv_entidad.add('B6:B100')

    dv_estado = DataValidation(type="list", formula1='"Pendiente,Pagado,Cancelado"')
    ws.add_data_validation(dv_estado)
    dv_estado.add('H6:H100')

    dv_cuenta = DataValidation(type="list", formula1='"' + ','.join(CONFIG['CUENTAS']) + '"')
    ws.add_data_validation(dv_cuenta)
    dv_cuenta.add('G6:G100')

    # Formato para filas de datos
    for row in range(6, 56):
        ws.cell(row=row, column=1).number_format = 'DD/MM/YYYY'
        ws.cell(row=row, column=6).number_format = '#,##0'
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border

    # Congelar
    ws.freeze_panes = 'A6'

    return ws


def crear_hoja_movimiento(wb):
    """Crear hoja MOVIMIENTO mensual"""
    ws = wb.create_sheet("MOVIMIENTO", 5)

    # Anchos
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 10

    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = 'MOVIMIENTO MENSUAL'
    ws['A1'].font = Font(bold=True, size=14)

    ws['B2'] = 'MES:'
    ws['C2'] = 'Enero'
    ws['C2'].font = Font(bold=True)

    row = 3

    # FAMILIA
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = 'FAMILIA'
    estilo_header(ws, ws[f'A{row}'], COLORES['VERDE_HEADER'])

    row += 1
    headers = ['CONCEPTO', 'DIA', 'TIPO', 'REAL', 'ESTADO', 'PRESUP.', '%']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color=COLORES['GRIS_CLARO'], end_color=COLORES['GRIS_CLARO'], fill_type='solid')
        cell.border = thin_border

    row += 1

    # Ingresos
    ws[f'A{row}'] = 'INGRESOS FAMILIA'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')
    row += 1

    ingresos_mov = [
        ('Salario Marco', 30, 'Fijo/Mensual', 8500000),
        ('Salario Marco NeuroTEA', 30, 'Fijo/Mensual', 5000000),
        ('Viatico Marco', '', 'Variable', 500000),
        ('Animador Biblico', 15, 'Variable/Mensual', 200000),
        ('Honorarios Clara NT', '', 'Variable', 3000000),
    ]

    start_ing = row
    for concepto, dia, tipo, ppto in ingresos_mov:
        ws.cell(row=row, column=1, value=concepto).border = thin_border
        ws.cell(row=row, column=2, value=dia).border = thin_border
        ws.cell(row=row, column=3, value=tipo).border = thin_border
        ws.cell(row=row, column=4, value=ppto).border = thin_border
        formato_numero(ws.cell(row=row, column=4))
        ws.cell(row=row, column=5, value='Pendiente').border = thin_border
        ws.cell(row=row, column=6, value=ppto).border = thin_border
        formato_numero(ws.cell(row=row, column=6))
        ws.cell(row=row, column=7, value=f'=IF(F{row}=0,"-",D{row}/F{row})').border = thin_border
        ws.cell(row=row, column=7).number_format = '0%'
        row += 1

    ws[f'A{row}'] = 'TOTAL INGRESOS'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f'=SUM(D{start_ing}:D{row-1})'
    ws[f'F{row}'] = f'=SUM(F{start_ing}:F{row-1})'
    formato_numero(ws[f'D{row}'])
    formato_numero(ws[f'F{row}'])
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORES['VERDE_CLARO'], end_color=COLORES['VERDE_CLARO'], fill_type='solid')

    row += 1

    # Egresos - Gastos Fijos
    ws[f'A{row}'] = 'EGRESOS - GASTOS FIJOS'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORES['ROJO_CLARO'], end_color=COLORES['ROJO_CLARO'], fill_type='solid')
    row += 1

    gastos_mov = [
        ('Salario Lili', 30, 'Fijo/Mensual', 1800000),
        ('Salario Laura', 30, 'Fijo/Mensual', 1200000),
        ('Escuela Ninos', 10, 'Fijo/Mensual', 1600000),
        ('Robotica Ninos', 11, 'Fijo/Mensual', 400000),
        ('ANDE Casa', 15, 'Variable/Mensual', 400000),
        ('Expensa Casa', 10, 'Variable/Mensual', 300000),
        ('Na Luisa', 5, 'Fijo/Mensual', 1200000),
        ('Contadora Marco', 15, 'Fijo/Mensual', 350000),
    ]

    start_gf = row
    for concepto, dia, tipo, ppto in gastos_mov:
        ws.cell(row=row, column=1, value=concepto).border = thin_border
        ws.cell(row=row, column=2, value=dia).border = thin_border
        ws.cell(row=row, column=3, value=tipo).border = thin_border
        ws.cell(row=row, column=4, value=ppto).border = thin_border
        formato_numero(ws.cell(row=row, column=4))
        ws.cell(row=row, column=5, value='Pendiente').border = thin_border
        ws.cell(row=row, column=6, value=ppto).border = thin_border
        formato_numero(ws.cell(row=row, column=6))
        ws.cell(row=row, column=7, value=f'=IF(F{row}=0,"-",D{row}/F{row})').border = thin_border
        ws.cell(row=row, column=7).number_format = '0%'
        row += 1

    ws[f'A{row}'] = 'SUBTOTAL GASTOS FIJOS'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f'=SUM(D{start_gf}:D{row-1})'
    ws[f'F{row}'] = f'=SUM(F{start_gf}:F{row-1})'
    formato_numero(ws[f'D{row}'])
    formato_numero(ws[f'F{row}'])
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = thin_border

    row += 1

    # Cuotas
    ws[f'A{row}'] = 'EGRESOS - CUOTAS'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    cuotas_mov = [
        ('Cajubi Marco', 10, 'Fijo/Mensual', 200000),
        ('Mutual Marco', 10, 'Variable/Mensual', 150000),
        ('Cuota ITAU', 15, 'Fijo/Mensual', 391000),
        ('Auto Laura', 20, 'Fijo/Mensual', 2000000),
        ('Coop Univ Clara', 15, 'Fijo/Mensual', 800000),
        ('Coomecipar Clara', 15, 'Fijo/Mensual', 500000),
        ('Solar Prestamo 1', 20, 'Fijo/Mensual', 850000),
        ('Solar Prestamo 2', 20, 'Fijo/Mensual', 650000),
    ]

    start_cuotas = row
    for concepto, dia, tipo, ppto in cuotas_mov:
        ws.cell(row=row, column=1, value=concepto).border = thin_border
        ws.cell(row=row, column=2, value=dia).border = thin_border
        ws.cell(row=row, column=3, value=tipo).border = thin_border
        ws.cell(row=row, column=4, value=ppto).border = thin_border
        formato_numero(ws.cell(row=row, column=4))
        ws.cell(row=row, column=5, value='Pendiente').border = thin_border
        ws.cell(row=row, column=6, value=ppto).border = thin_border
        formato_numero(ws.cell(row=row, column=6))
        ws.cell(row=row, column=7, value=f'=IF(F{row}=0,"-",D{row}/F{row})').border = thin_border
        ws.cell(row=row, column=7).number_format = '0%'
        row += 1

    ws[f'A{row}'] = 'SUBTOTAL CUOTAS'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f'=SUM(D{start_cuotas}:D{row-1})'
    ws[f'F{row}'] = f'=SUM(F{start_cuotas}:F{row-1})'
    formato_numero(ws[f'D{row}'])
    formato_numero(ws[f'F{row}'])
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = thin_border

    row += 1

    # Variables
    ws[f'A{row}'] = 'EGRESOS - VARIABLES'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    variables_mov = [
        ('Supermercado', '', 'Variable', 2500000),
        ('Combustible', '', 'Variable', 800000),
        ('Salud', '', 'Variable', 500000),
        ('Recreacion', '', 'Variable', 400000),
        ('Varios', '', 'Variable', 300000),
    ]

    start_var = row
    for concepto, dia, tipo, ppto in variables_mov:
        ws.cell(row=row, column=1, value=concepto).border = thin_border
        ws.cell(row=row, column=2, value=dia).border = thin_border
        ws.cell(row=row, column=3, value=tipo).border = thin_border
        ws.cell(row=row, column=4, value=0).border = thin_border
        formato_numero(ws.cell(row=row, column=4))
        ws.cell(row=row, column=5, value='Pendiente').border = thin_border
        ws.cell(row=row, column=6, value=ppto).border = thin_border
        formato_numero(ws.cell(row=row, column=6))
        ws.cell(row=row, column=7, value=f'=IF(F{row}=0,"-",D{row}/F{row})').border = thin_border
        ws.cell(row=row, column=7).number_format = '0%'
        row += 1

    ws[f'A{row}'] = 'SUBTOTAL VARIABLES'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f'=SUM(D{start_var}:D{row-1})'
    ws[f'F{row}'] = f'=SUM(F{start_var}:F{row-1})'
    formato_numero(ws[f'D{row}'])
    formato_numero(ws[f'F{row}'])
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = thin_border

    # Congelar
    ws.freeze_panes = 'A5'

    return ws


# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================

def crear_excel():
    """Crear el archivo Excel completo"""
    print("=" * 60)
    print("Creando Control Financiero 2026 V4...")
    print("Réplica exacta del V9 del usuario")
    print("=" * 60)

    wb = Workbook()

    # Eliminar hoja default
    del wb['Sheet']

    print("  → CONFIGURACION...")
    crear_hoja_configuracion(wb)

    print("  → TABLERO...")
    crear_hoja_tablero(wb)

    print("  → PRESUPUESTO...")
    crear_hoja_presupuesto(wb)

    print("  → GASTOS FIJOS...")
    crear_hoja_gastos_fijos(wb)

    print("  → CARGA...")
    crear_hoja_carga(wb)

    print("  → MOVIMIENTO...")
    crear_hoja_movimiento(wb)

    # Guardar
    filename = "Control_Financiero_2026_V4.xlsx"
    wb.save(filename)

    print(f"\n✅ Archivo creado: {filename}")
    return filename


def verificar(filename):
    """Verificar el archivo"""
    print(f"\n📋 Verificando {filename}...")

    wb = openpyxl.load_workbook(filename)

    print(f"  Hojas: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"    • {name}: {ws.max_row} filas x {ws.max_column} columnas")

    wb.close()
    print("\n✅ Verificación OK")


if __name__ == "__main__":
    filename = crear_excel()
    verificar(filename)
